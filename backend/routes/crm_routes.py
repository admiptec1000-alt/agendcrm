from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorDatabase
from database import get_database
from auth import get_current_user
from models import (
    TicketCreate, TicketUpdate, MessageCreate, QuickResponseCreate,
    CampaignCreate, CampaignUpdate, FlowCreate, FlowUpdate, AIChatRequest, AIChatResponse,
    TicketStatus
)
import uuid
import io
import base64
import re
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Any
from emergentintegrations.llm.chat import LlmChat, UserMessage
import os
from pydantic import BaseModel
from counters import next_ticket_number
from clients_link import find_or_create_client_by_phone, normalize_phone

router = APIRouter(prefix="/crm", tags=["crm"])

def _user_can_view_all_tickets(user: dict) -> bool:
    """Returns True if the user is allowed to see EVERY ticket in the
    company. Otherwise the user only sees:
      (a) tickets currently assigned to them (`assigned_to == user.id`)
      (b) unassigned tickets in queues/connections they are allowed to
          view (when `allowed_queue_ids` / `connection_ids` are set on
          the user). These act as the public pool — when another user
          claims one (POST /tickets/{id}/claim) it disappears from
          everyone else's listing.
      (c) when `view_connection_tickets` permission is set, ALSO every
          ticket whose `connection_id` is in `user.connection_ids` —
          regardless of who claimed it. This is what an "ops shift
          leader" wants: see every conversation that flows through
          their assigned WhatsApp instances.
    The permission is granted by either:
      - role (super_admin / company_admin) — implicit
      - explicit `view_all_tickets` flag in `user.permissions` (per-user toggle)
      - wildcard `*` permission (admin-equivalent profile)
    """
    role = (user.get("role") or "").lower()
    if role in ("super_admin", "superadmin", "company_admin"):
        return True
    perms = user.get("permissions") or []
    return "*" in perms or "view_all_tickets" in perms


def _user_can_view_connection_tickets(user: dict) -> bool:
    """True if the user has the `view_connection_tickets` permission. Used
    by `_ticket_visibility_filter` to expand the visibility scope to ALL
    tickets bound to the user's allowed connections (including those
    already assigned to another agent). Implicit for admins / wildcard.
    """
    if _user_can_view_all_tickets(user):
        return True
    perms = user.get("permissions") or []
    return "view_connection_tickets" in perms


def _ticket_visibility_filter(user: dict) -> dict:
    """Mongo `$or` clause to enforce the visibility rules above. Returns an
    empty dict (no filter) for users with the view-all permission.

    For non-admin users the "public pool" of unassigned tickets is restricted
    by the user's `allowed_queue_ids` and `connection_ids`. If both are
    empty/unset the user falls back to the legacy behaviour (all unassigned
    aberto tickets) to avoid silently breaking existing tenants that haven't
    configured the new RBAC yet.
    """
    if _user_can_view_all_tickets(user):
        return {}
    uid = user["id"]
    allowed_queues = user.get("allowed_queue_ids") or []
    allowed_conns = user.get("connection_ids") or []

    # `view_connection_tickets` — operator sees every ticket bound to one
    # of their connections, even when another agent owns it. This is the
    # "ops/shift lead" scope: still NOT global view, but wide within the
    # set of WhatsApp instances they participate in.
    extra_visibility = []
    if _user_can_view_connection_tickets(user) and allowed_conns:
        extra_visibility.append({"connection_id": {"$in": allowed_conns}})

    unassigned_match = [
        {"assigned_to": None},
        {"assigned_to": {"$exists": False}},
        {"assigned_to": ""},
    ]
    # Build the pool clause: tickets that ARE unassigned AND match at
    # least one queue/connection the user has access to.
    pool_extra = []
    if allowed_queues:
        pool_extra.append({"queue_id": {"$in": allowed_queues}})
    if allowed_conns:
        pool_extra.append({"connection_id": {"$in": allowed_conns}})

    if pool_extra:
        pool_clause = {
            "$and": [
                {"$or": unassigned_match},
                {"$or": pool_extra} if len(pool_extra) > 1 else pool_extra[0],
            ]
        }
    else:
        # No queue/connection scoping configured — fall back to legacy.
        pool_clause = {"$or": unassigned_match}

    return {
        "$or": [
            {"assigned_to": uid},
            pool_clause,
            *extra_visibility,
        ]
    }


async def _ensure_user_can_use_connection(
    db: AsyncIOMotorDatabase,
    user: dict,
    conn_id: str,
    require_connected: bool = True,
) -> dict:
    """Validates that the user is allowed to use the given WhatsApp
    connection for opening/transferring a ticket. Rules:
      - connection must belong to the user's company
      - non-admin users: connection must be in `user.connection_ids`
        (when the list is non-empty). An empty list = legacy unrestricted
        access for backwards compatibility.
      - when `require_connected=True`, the instance must have
        `status == "connected"` (UI also filters but server is source of truth).
    Returns the connection document.
    """
    conn = await db.channel_connections.find_one(
        {"id": conn_id, "company_id": user["company_id"]},
        {"_id": 0, "id": 1, "status": 1, "name": 1, "provider": 1},
    )
    if not conn:
        raise HTTPException(status_code=404, detail="Conexao nao encontrada")
    role = (user.get("role") or "").lower()
    is_admin = role in ("super_admin", "superadmin", "company_admin") or "*" in (user.get("permissions") or [])
    if not is_admin:
        allowed = user.get("connection_ids") or []
        if allowed and conn_id not in allowed:
            raise HTTPException(
                status_code=403,
                detail="Voce nao tem acesso a essa conexao. Use Transferir para repassar para outro usuario.",
            )
    if require_connected and (conn.get("status") or "").lower() != "connected":
        raise HTTPException(
            status_code=400,
            detail=f'A conexao "{conn.get("name") or conn_id}" nao esta conectada no momento.',
        )
    return conn




# Tickets
@router.get("/tickets")
async def list_tickets(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
    status: str = None,
    assigned_to: str = None,
    channel: str = None,
    search: str = None,
    tab: str = None,
    # 2026-06-25 — Sidebar filters now go SERVER-side so counts and list
    # always agree even when the company has >1000 tickets (the client
    # used to filter the truncated page locally → "Tab=46 / list=0").
    connection_id: str = None,
    queue_id: str = None,
    tag: str = None,
    limit: int = 1000,
    offset: int = 0,
):
    query = {"company_id": user["company_id"]}
    if status:
        query["status"] = status
    if assigned_to:
        query["assigned_to"] = assigned_to
    if channel:
        query["channel"] = channel
    if connection_id:
        query["connection_id"] = connection_id
    if queue_id:
        query["queue_id"] = queue_id
    if tag:
        # Match either by tag name or by tag id (tickets store either
        # depending on how they were created).
        td = await db.tags.find_one({"company_id": user["company_id"], "name": tag}, {"_id": 0, "id": 1})
        tag_values = [tag] + ([td["id"]] if (td and td.get("id")) else [])
        query["tags"] = {"$in": tag_values}
    if search:
        query["$or"] = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"customer_phone": {"$regex": search, "$options": "i"}},
        ]
    # NOTE: the tab filter must mirror the counts in /tickets/counts. The
    # operator expects "Atendendo" = tickets with an owner, "Aguardando" =
    # tickets in the public pool without an owner. Earlier this branch
    # filtered by *status* values (aberto/pago/...) which produced an empty
    # "Aguardando" list even when the count badge said 185.
    if tab == "atendendo":
        query["status"] = {"$nin": ["fechado", "cancelado"]}
        query["assigned_to"] = {"$nin": [None, ""]}
        query.setdefault("channel", {"$ne": "whatsapp_group"})
    elif tab == "aguardando":
        query["status"] = {"$nin": ["fechado", "cancelado"]}
        # Tickets without an assignee, regardless of whether they sit on a
        # kanban column or not (operator request — Aguardando is "everyone
        # who hasn't been pulled yet"). `$or` is folded into `$and` below
        # if a search filter is also active.
        assigned_clause = {"$or": [
            {"assigned_to": None},
            {"assigned_to": {"$exists": False}},
            {"assigned_to": ""},
        ]}
        if "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, assigned_clause]
        else:
            query.update(assigned_clause)
        query.setdefault("channel", {"$ne": "whatsapp_group"})
    elif tab == "grupos":
        # The "Grupos" tab ONLY shows whatsapp group conversations. Without
        # this branch the query fell through with no channel filter at all,
        # so every ticket in the company appeared under it.
        query["status"] = {"$nin": ["fechado", "cancelado"]}
        query["channel"] = "whatsapp_group"
    elif tab == "encerrados":
        # 2026-06-23 — Operators asked for a tab where they can audit
        # tickets the auto-close swept, OR ones they fechado manualmente,
        # and reopen the ones that were closed by mistake. The list still
        # respects the visibility filter applied below, so an operator
        # only sees closed tickets that they would have been able to see
        # while still open.
        query["status"] = "fechado"

    # Visibility: non-admin users only see their own tickets + the
    # unassigned-aberto pool. Admins / view_all_tickets see everything.
    vis_filter = _ticket_visibility_filter(user)
    if vis_filter:
        # If a `$or` already exists (from `search`), combine via $and
        if "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, vis_filter]
        else:
            query.update(vis_filter)

    # Clamp the limit to keep responses bounded but allow paging through
    # the full collection for tenants like Incinera with 2.5k+ tickets.
    safe_limit = max(1, min(int(limit or 1000), 5000))
    safe_offset = max(0, int(offset or 0))
    # 2026-05-27 — PERF: $slice na projecao mantem APENAS a ultima mensagem
    # de cada ticket (para `getLastMessage` no card de Atendimento). Reduz
    # 100x+ o payload em tenants com muitas mensagens por ticket.
    cursor = (
        db.tickets
        .find(query, {"_id": 0, "messages": {"$slice": -1}})
        .sort("updated_at", -1)
        .skip(safe_offset)
        .limit(safe_limit)
    )
    tickets = await cursor.to_list(safe_limit)
    # Annotate each ticket with the registered client name (when the phone
    # matches a record in `clients`). The list UI shows this instead of the
    # raw WhatsApp pushName so the operator sees the CRM-canonical name.
    phones = list({t.get("customer_phone") for t in tickets if t.get("customer_phone")})
    if phones:
        clients = await db.clients.find(
            {"company_id": user["company_id"], "phone": {"$in": phones}},
            {"_id": 0, "phone": 1, "name": 1},
        ).to_list(len(phones))
        name_by_phone = {c["phone"]: c.get("name") for c in clients if c.get("name")}
        for t in tickets:
            ph = t.get("customer_phone")
            if ph and name_by_phone.get(ph):
                t["client_registered_name"] = name_by_phone[ph]
    return tickets

@router.get("/tickets/counts")
async def get_ticket_counts(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
    channel: str = None,
    search: str = None,
    queue_id: str = None,
    connection_id: str = None,
    assigned_to: str = None,
    tag: str = None,
):
    """2026-06-24 — Contadores agora respeitam os MESMOS filtros que a
    listagem em /tickets (channel, search, queue_id, connection_id,
    assigned_to, tag). Antes os badges mostravam totais brutos
    (ex.: Aguardando=48) enquanto a lista filtrada exibia 0 resultados
    → operador confundido. Agora os numeros refletem fielmente a
    quantidade real visivel naquele perfil com aqueles filtros."""
    company_id = user["company_id"]
    base = {"company_id": company_id}
    vis = _ticket_visibility_filter(user)
    if vis:
        base = {**base, **vis}
    # Apply listing-level filters so the counts always match the rows
    # the operator can actually see.
    if channel:
        base["channel"] = channel
    if queue_id:
        base["queue_id"] = queue_id
    if connection_id:
        base["connection_id"] = connection_id
    if assigned_to:
        base["assigned_to"] = assigned_to
    if tag:
        td = await db.tags.find_one({"company_id": company_id, "name": tag}, {"_id": 0, "id": 1})
        tag_values = [tag] + ([td["id"]] if (td and td.get("id")) else [])
        base["tags"] = {"$in": tag_values}
    if search:
        rx = {"$regex": search, "$options": "i"}
        base["$or"] = [
            {"customer_name": rx},
            {"customer_phone": rx},
            {"description": rx},
            {"messages.content": rx},
        ]
    atendendo = await db.tickets.count_documents({**base, "status": {"$nin": ["fechado", "cancelado"]}, "assigned_to": {"$nin": [None, ""]}, "channel": {"$ne": "whatsapp_group"}} if not channel else {**base, "status": {"$nin": ["fechado", "cancelado"]}, "assigned_to": {"$nin": [None, ""]}})
    aguardando_q = {
        **base, "status": {"$nin": ["fechado", "cancelado"]},
        "$or": [{"assigned_to": None}, {"assigned_to": {"$exists": False}}, {"assigned_to": ""}],
    }
    if not channel:
        aguardando_q["channel"] = {"$ne": "whatsapp_group"}
    # When `search` is set, base already has $or for text — adding another
    # $or would break Mongo. Wrap previous $or under $and for combination.
    if "$or" in base and base.get("$or") != aguardando_q.get("$or"):
        # Two distinct $or clauses → combine via $and
        aguardando_q = {k: v for k, v in aguardando_q.items() if k != "$or"}
        aguardando_q["$and"] = [
            {"$or": base["$or"]},
            {"$or": [{"assigned_to": None}, {"assigned_to": {"$exists": False}}, {"assigned_to": ""}]},
        ]
    aguardando = await db.tickets.count_documents(aguardando_q)
    grupos = await db.tickets.count_documents({**base, "status": {"$nin": ["fechado", "cancelado"]}, "channel": "whatsapp_group"})
    encerrados = await db.tickets.count_documents({**base, "status": "fechado"})
    total = await db.tickets.count_documents(base)
    # 2026-02-28 — Adicionado contador "fechados_hoje" pro card da
    # tela Inicio. Usa `closed_at` quando disponivel, senao `updated_at`.
    from datetime import datetime as _dt, timezone as _tz
    today_prefix = _dt.now(_tz.utc).date().isoformat()
    fechados_hoje = await db.tickets.count_documents({
        **base,
        "status": {"$in": ["fechado", "closed"]},
        "$or": [
            {"closed_at": {"$regex": f"^{today_prefix}"}},
            {"closed_at": {"$exists": False}, "updated_at": {"$regex": f"^{today_prefix}"}},
        ],
    })
    return {"atendendo": atendendo, "aguardando": aguardando, "grupos": grupos, "encerrados": encerrados, "total": total, "fechados_hoje": fechados_hoje}


@router.post("/tickets/open-for-client")
async def open_ticket_for_client(
    payload: dict,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Find-or-create an OPEN WhatsApp ticket for the given client.

    Used by the "Abrir atendimento" shortcut on the Clientes page so the
    operator can jump straight to the chat without searching by hand.
    The caller passes `{ client_id, phone, name }`. We:

      1) Look for an open ticket (`status not in [fechado, cancelado]`,
         not a group) bound either to the client_id or the phone.
      2) If none exists, create a new whatsapp ticket bound to the
         company's first connected WhatsApp connection. If no connection
         is connected, we still create the ticket so the operator can
         enqueue a message; sending will be queued by the existing logic.
      3) Return the full ticket so the frontend can navigate to it.
    """
    company_id = user["company_id"]
    client_id = payload.get("client_id")
    phone = (payload.get("phone") or "").strip()
    name = (payload.get("name") or phone).strip()
    if not phone and not client_id:
        raise HTTPException(400, "client_id ou phone obrigatorio")

    digits_only = re.sub(r"\D", "", phone) if phone else ""

    # Find an open ticket by client_id OR by phone (digits-only OR raw).
    candidates_or = []
    if client_id:
        candidates_or.append({"client_id": client_id})
    if digits_only:
        candidates_or.append({"customer_phone": digits_only})
    if phone and phone != digits_only:
        candidates_or.append({"customer_phone": phone})
    ticket = None
    if candidates_or:
        ticket = await db.tickets.find_one(
            {
                "company_id": company_id,
                "status": {"$nin": ["fechado", "cancelado"]},
                "channel": {"$ne": "whatsapp_group"},
                "$or": candidates_or,
            },
            {"_id": 0},
        )

    if not ticket:
        # Pick first connected WhatsApp connection (or first one if none
        # are connected) so the new ticket has a sensible default channel.
        conn = await db.channel_connections.find_one(
            {"company_id": company_id, "status": "connected"}, {"_id": 0}
        ) or await db.channel_connections.find_one(
            {"company_id": company_id}, {"_id": 0}
        )
        new_id = str(uuid.uuid4())
        from counters import next_ticket_number as _next
        n = await _next(db, company_id)
        ticket = {
            "id": new_id,
            "ticket_number": n,
            "company_id": company_id,
            "connection_id": (conn or {}).get("id"),
            "client_id": client_id,
            "customer_name": name,
            "customer_phone": digits_only or phone,
            "customer_email": None,
            "status": "aberto",
            "priority": "medium",
            "channel": "whatsapp",
            "is_group": False,
            "description": None,
            "assigned_to": user["id"],
            "queue_id": ((conn or {}).get("queue_ids") or [None])[0] if len((conn or {}).get("queue_ids") or []) == 1 else None,
            "messages": [],
            "tags": [],
            "value": 0.0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "origin": "client_shortcut",
        }
        await db.tickets.insert_one(ticket)
        ticket.pop("_id", None)
    return ticket


@router.get("/tickets/{ticket_id}")
async def get_ticket(
    ticket_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    ticket = await db.tickets.find_one(
        {"id": ticket_id, "company_id": user["company_id"]}, {"_id": 0}
    )
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")
    # Mark this ticket as read for the current user. The unread badge in
    # the sidebar uses `ticket.read_state[user_id]` to know which inbound
    # messages are still pending — opening the conversation resets the
    # counter immediately.
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.tickets.update_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"$set": {f"read_state.{user['id']}": now_iso}},
    )
    ticket.setdefault("read_state", {})[user["id"]] = now_iso
    # Synthesize media_url / media_kind / media_mimetype on outbound messages
    # so the frontend renders the playable bubble for audios/images/videos
    # sent FROM the platform (not only inbound ones from the webhook).
    # The actual bytes live in `attachment_data_b64`; we expose them via
    # the streaming endpoint below to keep the JSON light.
    for m in (ticket.get("messages") or []):
        if m.get("attachment_kind") and m.get("attachment_data_b64") and not m.get("media_url"):
            m["media_url"] = f"/api/crm/tickets/{ticket_id}/messages/{m.get('id')}/attachment"
            m["media_kind"] = m.get("attachment_kind")
            m["media_mimetype"] = m.get("attachment_mimetype")
            m["media_filename"] = m.get("attachment_filename")
        # Drop the heavy base64 blob from the JSON response.
        m.pop("attachment_data_b64", None)
    return ticket


@router.get("/tickets/{ticket_id}/messages/{message_id}/attachment")
async def stream_message_attachment(
    ticket_id: str,
    message_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Stream the binary stored in `attachment_data_b64` so <audio>, <img>,
    <video> and <a> tags can fetch the asset by URL instead of paying the
    base64 inflation cost in the ticket JSON.

    Audio uploads from the operator (PTT recordings) are persisted as
    base64 inline on the message — without this endpoint the chat would
    never render a player for them (only inbound ones get a `media_url`).
    """
    ticket = await db.tickets.find_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"_id": 0, "messages": 1},
    )
    if not ticket:
        raise HTTPException(404, "Ticket nao encontrado")
    msg = next((m for m in (ticket.get("messages") or []) if m.get("id") == message_id), None)
    if not msg or not msg.get("attachment_data_b64"):
        raise HTTPException(404, "Anexo nao encontrado")
    try:
        raw = base64.b64decode(msg["attachment_data_b64"])
    except Exception:
        raise HTTPException(500, "Anexo corrompido")
    mime = msg.get("attachment_mimetype") or "application/octet-stream"
    headers = {}
    if msg.get("attachment_filename"):
        headers["Content-Disposition"] = f'inline; filename="{msg["attachment_filename"]}"'
    return Response(content=raw, media_type=mime, headers=headers)

@router.post("/tickets")
async def create_ticket(
    data: TicketCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    # Duplicate guard: refuse to open a second OPEN ticket for the same

    # phone in the same tenant unless the operator explicitly forces it.
    # The match is digits-only so "5511999..." and "(55) 11 999..." are
    # treated as the same person. Group tickets are excluded — they're
    # keyed by group_jid, not by individual phone.
    digits_phone = re.sub(r"\D", "", data.customer_phone or "")
    if digits_phone and not data.force_create:
        candidates_or = [{"customer_phone": digits_phone}]
        if data.customer_phone and data.customer_phone != digits_phone:
            candidates_or.append({"customer_phone": data.customer_phone})
        existing = await db.tickets.find_one(
            {
                "company_id": user["company_id"],
                "status": {"$nin": ["fechado", "cancelado"]},
                "channel": {"$ne": "whatsapp_group"},
                "$or": candidates_or,
            },
            {"_id": 0, "id": 1, "ticket_number": 1, "customer_name": 1,
             "customer_phone": 1, "status": 1, "assigned_to": 1,
             "updated_at": 1},
        )
        if existing:
            # 409 Conflict — frontend reads `detail.existing_ticket` to offer
            # "open existing" vs. "create anyway".
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "duplicate_open_ticket",
                    "message": f"Já existe um atendimento aberto (#{existing.get('ticket_number')}) para o telefone {data.customer_phone}.",
                    "existing_ticket": existing,
                },
            )

    # Validate the connection the operator picked: it must belong to the
    # company, be `connected`, AND be one of the user's allowed_connections
    # (when the user has any restriction). Company admins bypass the
    # whitelist but still need a connected instance.
    conn_doc = None
    if data.connection_id:
        conn_doc = await _ensure_user_can_use_connection(
            db, user, data.connection_id, require_connected=True
        )

    ticket_id = str(uuid.uuid4())
    ticket_number = await next_ticket_number(db, user["company_id"])
    client_id = await find_or_create_client_by_phone(
        db, user["company_id"], data.customer_phone,
        name=data.customer_name, email=data.customer_email
    )
    ticket = {
        "id": ticket_id,
        "ticket_number": ticket_number,
        "company_id": user["company_id"],
        "client_id": client_id,
        "customer_name": data.customer_name,
        "customer_phone": data.customer_phone,
        "customer_email": data.customer_email,
        "status": data.status,
        "priority": data.priority,
        "channel": data.channel,
        "connection_id": (conn_doc or {}).get("id"),
        "description": data.description,
        "assigned_to": None,
        "messages": [],
        "tags": data.tags or [],
        "value": float(data.value or 0),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.tickets.insert_one(ticket)
    return {k: v for k, v in ticket.items() if k != "_id"}

@router.put("/tickets/{ticket_id}")
async def update_ticket(
    ticket_id: str,
    data: TicketUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket não encontrado")

    raw = data.model_dump(exclude_unset=True)
    # Fields where an explicit null means "clear this value" (not "ignore").
    CLEARABLE_FIELDS = {"kanban_column_id", "queue_id", "connection_id", "assigned_to"}
    update_data = {
        k: v for k, v in raw.items()
        if v is not None or k in CLEARABLE_FIELDS
    }
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    # If the operator is switching the ticket to a different WhatsApp
    # instance from the chat header, enforce the same access rules used
    # at creation time. Clearing the connection (explicit null) is
    # allowed without validation — that just detaches the ticket.
    if "connection_id" in update_data and update_data["connection_id"]:
        new_conn_id = update_data["connection_id"]
        if new_conn_id != ticket.get("connection_id"):
            await _ensure_user_can_use_connection(
                db, user, new_conn_id, require_connected=True
            )

    # When the ticket is being closed/cancelled, also clear the bot_paused
    # flag so that if the customer comes back later and a fresh ticket is
    # opened, the bot can run normally. We do NOT auto-resume on the same
    # ticket — only on a new ticket lifecycle.
    new_status = update_data.get("status")
    if new_status in ("fechado", "cancelado") and ticket.get("bot_paused"):
        update_data["bot_paused"] = False
        update_data["bot_paused_at"] = None
        update_data["bot_paused_reason"] = None

    if update_data:
        await db.tickets.update_one(
            {"id": ticket_id},
            {"$set": update_data}
        )

    # 2026-05-28 — Quando o operador fecha MANUALMENTE o ticket, opcional
    # mente envia a MESMA mensagem de encerramento usada pelo auto-close
    # (companies.ticket_auto_close_message). Controlado por
    # `companies.send_close_message_on_manual` (default False). Best-effort,
    # falha NAO derruba a request.
    if new_status == "fechado":
        try:
            comp = await db.companies.find_one(
                {"id": user["company_id"]},
                {"_id": 0, "name": 1, "send_close_message_on_manual": 1, "ticket_auto_close_message": 1},
            ) or {}
            if bool(comp.get("send_close_message_on_manual")) and (comp.get("ticket_auto_close_message") or "").strip():
                import httpx, os, logging
                _logger = logging.getLogger(__name__)
                phone = ticket.get("customer_phone") or ""
                connection_id = ticket.get("connection_id") or ticket.get("channel_id")
                contact_name = ticket.get("customer_name") or ""
                if phone and connection_id:
                    company_name = comp.get("name") or ""
                    # 2026-02-28 — Mesma logica SGP do scheduler.
                    fvars = ticket.get("flow_vars") or {}
                    nome_sgp = (fvars.get("nome_cliente") or "").strip()
                    primeiro_nome_sgp = nome_sgp.split()[0] if nome_sgp else ""
                    template = comp["ticket_auto_close_message"]
                    msg = (
                        template
                        .replace("{{nome}}", contact_name)
                        .replace("{nome}", contact_name)
                        .replace("{{empresa}}", company_name)
                        .replace("{empresa}", company_name)
                        .replace("{{nome_sgp}}", nome_sgp)
                        .replace("{nome_sgp}", nome_sgp)
                        .replace("{{primeiro_nome_sgp}}", primeiro_nome_sgp)
                        .replace("{primeiro_nome_sgp}", primeiro_nome_sgp)
                    )
                    wa_url = os.environ.get("WA_SERVICE_URL", "http://localhost:3002")
                    try:
                        async with httpx.AsyncClient(timeout=10.0) as client:
                            r = await client.post(
                                f"{wa_url}/instances/{connection_id}/send",
                                json={"phone": phone, "message": msg},
                            )
                            _logger.info(
                                f"[manual-close] msg sent ticket={ticket_id} status={r.status_code}"
                            )
                        await db.tickets.update_one(
                            {"id": ticket_id},
                            {"$push": {"messages": {
                                "from": "bot",
                                "text": msg,
                                "type": "text",
                                "timestamp": datetime.now(timezone.utc).isoformat(),
                                "system": True,
                                "reason": "manual_close",
                            }}},
                        )
                    except Exception as se:
                        _logger.warning(f"[manual-close] msg failed ticket={ticket_id}: {se}")
        except Exception as ce:
            import logging
            logging.getLogger(__name__).warning(f"[manual-close] config lookup failed ticket={ticket_id}: {ce}")

    updated_ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    return updated_ticket

@router.delete("/tickets/{ticket_id}")
async def delete_ticket(
    ticket_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    result = await db.tickets.delete_one({"id": ticket_id, "company_id": user["company_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")
    return {"deleted": True}


@router.post("/tickets/{src_id}/merge-into/{dst_id}")
async def merge_tickets(
    src_id: str,
    dst_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    """Merges the SOURCE ticket into the DESTINATION ticket and deletes the source.

    Useful when WhatsApp Linked Devices (@lid) create a duplicate ticket with a
    fake phone number. The admin identifies the real ticket, picks the merge
    action on the duplicate, and this endpoint:

      1. Appends all messages from src into dst (deduped by wa_message_id)
      2. Carries over tags that dst doesn't already have
      3. Preserves dst's ticket_number / kanban_column_id / client_id
      4. Deletes the src ticket

    The destination always wins on single-valued fields; only messages/tags
    are additive. Multi-tenant safe (both must belong to the caller's company).
    """
    if src_id == dst_id:
        raise HTTPException(400, "Ticket de origem e destino sao o mesmo")
    company_id = user["company_id"]
    src = await db.tickets.find_one({"id": src_id, "company_id": company_id})
    dst = await db.tickets.find_one({"id": dst_id, "company_id": company_id})
    if not src or not dst:
        raise HTTPException(404, "Ticket de origem ou destino nao encontrado")

    src_msgs = src.get("messages") or []
    existing_wa_ids = {m.get("wa_message_id") for m in (dst.get("messages") or []) if m.get("wa_message_id")}
    new_msgs = [m for m in src_msgs if not (m.get("wa_message_id") and m.get("wa_message_id") in existing_wa_ids)]

    # Merge tags (unique by tag id/name)
    dst_tags = dst.get("tags") or []
    dst_tag_keys = {(t.get("id") or t.get("name")) for t in dst_tags}
    new_tags = [t for t in (src.get("tags") or []) if (t.get("id") or t.get("name")) not in dst_tag_keys]

    update = {}
    if new_msgs:
        update["$push"] = {"messages": {"$each": new_msgs}}
    set_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if new_tags:
        # Use $addToSet style via combined array to preserve ordering
        set_fields["tags"] = dst_tags + new_tags
    update["$set"] = set_fields

    await db.tickets.update_one({"id": dst_id, "company_id": company_id}, update)
    await db.tickets.delete_one({"id": src_id, "company_id": company_id})

    # Re-point any quotes that were attached to the src ticket to dst
    await db.quotes.update_many(
        {"ticket_id": src_id, "company_id": company_id},
        {"$set": {"ticket_id": dst_id}}
    )

    return {
        "merged": True,
        "into_ticket_id": dst_id,
        "into_ticket_number": dst.get("ticket_number"),
        "messages_added": len(new_msgs),
        "tags_added": len(new_tags),
    }


class ResolveLidRequest(BaseModel):
    real_phone: str


@router.post("/tickets/{ticket_id}/claim")
async def claim_ticket(
    ticket_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Operator pulls (puxa) a ticket from the unassigned-aberto pool.
    From this moment on, the ticket disappears from EVERY other operator's
    list (unless they have `view_all_tickets`). Idempotent: re-claiming
    a ticket already assigned to the caller is a no-op success."""
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(404, "Ticket nao encontrado")
    current = ticket.get("assigned_to")
    if current and current != user["id"] and not _user_can_view_all_tickets(user):
        # Another operator already owns it AND we don't have view-all rights.
        raise HTTPException(409, "Atendimento ja foi puxado por outro operador")
    await db.tickets.update_one(
        {"id": ticket_id},
        {"$set": {"assigned_to": user["id"], "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return await db.tickets.find_one({"id": ticket_id}, {"_id": 0})


@router.post("/tickets/{ticket_id}/release")
async def release_ticket(
    ticket_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Returns a claimed ticket back to the public pool (sets `assigned_to`
    to null). Only the current owner OR an admin can release. Used when
    the operator realises another colleague is better-suited for the case."""
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(404, "Ticket nao encontrado")
    current = ticket.get("assigned_to")
    if current and current != user["id"] and not _user_can_view_all_tickets(user):
        raise HTTPException(403, "Apenas o atendente atual ou um admin pode liberar este atendimento")
    await db.tickets.update_one(
        {"id": ticket_id},
        {"$set": {"assigned_to": None, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return await db.tickets.find_one({"id": ticket_id}, {"_id": 0})


@router.post("/tickets/{ticket_id}/reopen")
async def reopen_ticket(
    ticket_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Reopens a closed (`fechado`) ticket. Clears the close metadata
    (`closed_at`, `closed_reason`) and flips status back to `aberto` so it
    surfaces again in the regular Atendimentos listings. The operator
    needs the same visibility rights they would for any other ticket of
    that connection — i.e., either it was assigned to them, or they have
    `view_all_tickets` / `view_connection_tickets`, or it lives in a
    connection/queue they participate in.
    """
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(404, "Ticket nao encontrado")
    if (ticket.get("status") or "") != "fechado":
        raise HTTPException(400, "O ticket nao esta fechado")
    # Apply the same visibility guard so operators dont reopen tickets
    # they shouldn't even see.
    vis = _ticket_visibility_filter(user)
    if vis:
        in_scope = await db.tickets.find_one(
            {"id": ticket_id, "company_id": user["company_id"], **vis},
            {"_id": 0, "id": 1},
        )
        if not in_scope:
            raise HTTPException(403, "Voce nao tem permissao para reabrir esse atendimento")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.tickets.update_one(
        {"id": ticket_id},
        {
            "$set": {"status": "aberto", "updated_at": now_iso, "reopened_at": now_iso, "reopened_by": user["id"]},
            "$unset": {"closed_at": "", "closed_reason": ""},
        },
    )
    return await db.tickets.find_one({"id": ticket_id}, {"_id": 0})




@router.post("/tickets/{ticket_id}/resolve-lid")
async def resolve_ticket_lid(
    ticket_id: str,
    body: ResolveLidRequest,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Manual fallback for hidden-number contacts (`pending_lid_resolution=True`).

    The operator obtains the real phone via voice/email/business card and
    submits it here. Same merge logic as `/channels/webhook/lid-resolved`:
    if another open ticket already exists for the real phone, messages are
    merged into that one and the LID-only ticket is deleted; otherwise the
    LID ticket is promoted to the real phone in place.
    """
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(404, "Ticket nao encontrado")
    if not ticket.get("lid_jid"):
        raise HTTPException(400, "Este ticket nao tem numero oculto pendente")
    # Reuse the same logic from the webhook handler so behavior is identical
    from routes.channels_routes import _apply_lid_resolution
    result = await _apply_lid_resolution(db, user["company_id"], ticket["lid_jid"], body.real_phone)
    if not result.get("updated"):
        raise HTTPException(400, f"Nao foi possivel resolver: {result.get('reason')}")
    return result


@router.get("/clients/{client_id}/timeline")
async def get_client_timeline(
    client_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
    limit: int = 50,
):
    """360° view of a client: full ticket history + summary stats.

    Returns tickets ordered by created_at desc plus aggregated stats so the
    chat sidebar can render a single panel without N+1 queries.
    """
    company_id = user["company_id"]

    # Verify the client belongs to this tenant
    client = await db.clients.find_one(
        {"id": client_id, "company_id": company_id}, {"_id": 0}
    )
    if not client:
        raise HTTPException(status_code=404, detail="Cliente nao encontrado")

    cursor = db.tickets.find(
        {"company_id": company_id, "client_id": client_id},
        {"_id": 0, "id": 1, "ticket_number": 1, "status": 1, "value": 1,
         "channel": 1, "created_at": 1, "closed_at": 1, "updated_at": 1,
         "customer_name": 1, "tags": 1, "rating": 1, "kanban_column_id": 1},
    ).sort("created_at", -1).limit(limit)
    tickets = await cursor.to_list(limit)

    # Stats are computed over the FULL history (independent from the paginated
    # tickets array) so values stay correct on high-volume clients.
    pipeline = [
        {"$match": {"company_id": company_id, "client_id": client_id}},
        {"$group": {
            "_id": None,
            "total_tickets": {"$sum": 1},
            "open": {"$sum": {"$cond": [
                {"$not": {"$in": ["$status", ["fechado", "cancelado"]]}}, 1, 0
            ]}},
            "closed": {"$sum": {"$cond": [{"$eq": ["$status", "fechado"]}, 1, 0]}},
            "total_value": {"$sum": {"$ifNull": ["$value", 0]}},
            "last_visit": {"$max": "$created_at"},
        }},
    ]
    agg = await db.tickets.aggregate(pipeline).to_list(1)
    s = agg[0] if agg else {"total_tickets": 0, "open": 0, "closed": 0, "total_value": 0, "last_visit": None}
    total = s.get("total_tickets") or 0
    total_value = float(s.get("total_value") or 0)
    avg_value = (total_value / total) if total else 0.0

    return {
        "client": client,
        "stats": {
            "total_tickets": total,
            "open": s.get("open") or 0,
            "closed": s.get("closed") or 0,
            "total_value": total_value,
            "avg_value": avg_value,
            "last_visit": s.get("last_visit"),
        },
        "tickets": tickets,
    }



@router.get("/tickets/{ticket_id}/client")
async def get_ticket_client(
    ticket_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    """Returns the linked client (cliente/lead). When the ticket has no
    client_id yet (legacy), tries to match by phone and links it lazily.
    """
    ticket = await db.tickets.find_one(
        {"id": ticket_id, "company_id": user["company_id"]}, {"_id": 0}
    )
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")

    cid = ticket.get("client_id")
    if not cid:
        cid = await find_or_create_client_by_phone(
            db, user["company_id"],
            ticket.get("customer_phone", ""),
            name=ticket.get("customer_name"),
            email=ticket.get("customer_email"),
        )
        if cid:
            await db.tickets.update_one({"id": ticket_id}, {"$set": {"client_id": cid}})

    if not cid:
        # No phone, no lead yet — return a stub the frontend can fill.
        return {
            "id": None,
            "name": ticket.get("customer_name", ""),
            "phone": ticket.get("customer_phone", ""),
            "email": ticket.get("customer_email", ""),
            "person_type": "fisica",
        }
    client = await db.clients.find_one({"id": cid, "company_id": user["company_id"]}, {"_id": 0})
    return client or {"id": cid}


@router.put("/tickets/{ticket_id}/client")
async def update_ticket_client(
    ticket_id: str,
    payload: dict,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    """Updates (or creates) the linked client and refreshes denormalized
    customer_* fields on the ticket so the chat header stays in sync.
    """
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")

    cid = ticket.get("client_id")
    company_id = user["company_id"]

    # Whitelist the fields a chat operator can edit on the linked client
    ALLOWED = {
        "name", "phone", "email", "person_type", "cpf", "cnpj",
        "company_name", "cep", "address", "city", "state",
        "birth_date", "notes",
    }
    fields = {k: v for k, v in (payload or {}).items() if k in ALLOWED}

    if not cid:
        # First save creates the client document
        if not fields.get("phone"):
            fields["phone"] = ticket.get("customer_phone", "")
        if not fields.get("name"):
            fields["name"] = ticket.get("customer_name") or fields.get("phone") or "Cliente"
        cid = str(uuid.uuid4())
        await db.clients.insert_one({
            "id": cid, "company_id": company_id,
            "person_type": fields.get("person_type") or "fisica",
            "total_appointments": 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_via": "ticket_panel",
            **fields,
        })
    else:
        if fields:
            await db.clients.update_one(
                {"id": cid, "company_id": company_id}, {"$set": fields}
            )

    # Sync denormalized fields on the ticket so the chat header stays correct.
    sync = {"updated_at": datetime.now(timezone.utc).isoformat(), "client_id": cid}
    if "name" in fields:
        sync["customer_name"] = fields["name"]
    if "phone" in fields:
        sync["customer_phone"] = fields["phone"]
    if "email" in fields:
        sync["customer_email"] = fields["email"]
    await db.tickets.update_one({"id": ticket_id}, {"$set": sync})

    client = await db.clients.find_one({"id": cid, "company_id": company_id}, {"_id": 0})
    return client

@router.post("/tickets/{ticket_id}/messages")
async def add_message_to_ticket(
    ticket_id: str,
    data: MessageCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket não encontrado")

    # Signature prefix (operator name) — opt-in via MessageCreate.with_signature.
    # Default True so existing behavior is preserved when callers omit the flag.
    outbound_content = data.content
    if (
        data.sender_type == "agent"
        and getattr(data, "with_signature", True)
        and ticket.get("channel") == "whatsapp"
        and (user.get("name") or "").strip()
    ):
        sig = (user["name"] or "").strip()
        outbound_content = f"*{sig}:*\n{data.content}"

    message = {
        "id": str(uuid.uuid4()),
        "content": outbound_content,
        "sender_type": data.sender_type,
        "sender_id": user["id"],
        "sender_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "delivery_status": "pending",
    }

    # If agent and channel is whatsapp, actually send via Baileys
    delivery_error = None
    if data.sender_type == "agent" and ticket.get("channel") == "whatsapp" and ticket.get("customer_phone"):
        try:
            import httpx
            import os as _os
            import logging as _lg
            _log = _lg.getLogger(__name__)
            wa_url = _os.environ.get("WA_SERVICE_URL", "http://localhost:3002")
            # Prefer ticket's bound connection but VALIDATE it is currently
            # connected — old tickets sometimes point to a rotated/deleted
            # connection_id and silently fail. Fall back to any healthy
            # WhatsApp connection in the company.
            conn_id = ticket.get("connection_id")
            if conn_id:
                _conn_doc = await db.channel_connections.find_one(
                    {"id": conn_id, "company_id": user["company_id"]},
                    {"_id": 0, "id": 1, "status": 1},
                )
                if not _conn_doc or (_conn_doc.get("status") or "").lower() != "connected":
                    conn_id = None  # force fallback below
            if not conn_id:
                conn = await db.channel_connections.find_one(
                    {"company_id": user["company_id"], "type": "whatsapp", "status": "connected"},
                    {"_id": 0, "id": 1}
                )
                conn_id = conn["id"] if conn else None
            # Hidden-number contacts (LID-only): WhatsApp won't accept the
            # raw LID digits via onWhatsApp() — the only address that works
            # is the original `XXX@lid` JID. The microservice's /send
            # endpoint already passes through any value containing `@`
            # straight to sendMessage (no JID resolution), so we just send
            # the lid_jid as `phone`.
            target_phone = ticket["customer_phone"]
            if ticket.get("pending_lid_resolution") and ticket.get("lid_jid"):
                target_phone = ticket["lid_jid"]
            if conn_id:
                async with httpx.AsyncClient(timeout=30.0) as client:
                    resp = await client.post(
                        f"{wa_url}/instances/{conn_id}/send",
                        json={"phone": target_phone, "message": outbound_content}
                    )
                    try:
                        res = resp.json()
                    except Exception:
                        res = {}
                    if resp.status_code == 200 and res.get("success"):
                        message["delivery_status"] = "sent"
                        message["wa_message_id"] = res.get("jid")
                    else:
                        message["delivery_status"] = "failed"
                        delivery_error = res.get("error") or f"HTTP {resp.status_code}: {resp.text[:80]}"
                        _log.warning(f"WA send failed for ticket {ticket_id}: {delivery_error}")
            else:
                message["delivery_status"] = "failed"
                delivery_error = "Nenhuma conexao WhatsApp ativa"
        except Exception as e:
            message["delivery_status"] = "failed"
            delivery_error = str(e)[:200]

    if delivery_error:
        message["delivery_error"] = delivery_error

    update_set = {"updated_at": datetime.now(timezone.utc).isoformat()}
    # Track last outgoing for the @lid fallback in the webhook handler.
    # When an incoming message arrives later with a fake LID phone, the
    # webhook resolves the destination ticket by looking up the most
    # recent outgoing on the same company (5-min window).
    if (
        data.sender_type == "agent"
        and ticket.get("channel") == "whatsapp"
        and message.get("delivery_status") == "sent"
    ):
        update_set["last_outgoing_at"] = update_set["updated_at"]
        # Bind the connection_id to the ticket if it wasn't already
        # (manual-created tickets often start without one). The ticket
        # remembers which connection it lives on, helping the webhook
        # match incoming replies even when the user replies via @lid.
        if not ticket.get("connection_id") and conn_id:
            update_set["connection_id"] = conn_id

    await db.tickets.update_one(
        {"id": ticket_id},
        {"$push": {"messages": message}, "$set": update_set}
    )

    # Pause the bot for this ticket if the company opted-in. Only relevant
    # when the operator sent the message (sender_type=agent) AND there's an
    # active flow waiting. See bot_pause.pause_bot_on_ticket_if_enabled.
    if data.sender_type == "agent":
        try:
            from bot_pause import pause_bot_on_ticket_if_enabled
            await pause_bot_on_ticket_if_enabled(
                db, ticket, reason="agent_message_platform"
            )
        except Exception as e:
            import logging as _lg
            _lg.getLogger(__name__).warning(f"[bot_pause] platform-send failed: {e}")

    return message


# === TICKET TAGS ===
class TicketTagToggle(BaseModel):
    tag: str


@router.post("/tickets/{ticket_id}/media")
async def send_media_to_ticket(
    ticket_id: str,
    payload: dict,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Send a media file (image/audio/video/document) to the customer via WA.

    Body: {
      filename: str,
      mimetype: str,
      data_base64: str,
      caption: str | None,    # optional text shown with the file (not for audio)
    }

    Persists the message on the ticket and forwards the bytes to the WA
    microservice. Audio with `audio/*` mimetype is sent as a PTT voice note.
    """
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")
    # Accept any WhatsApp channel — `whatsapp` (private chat) AND
    # `whatsapp_group`. Previously only `whatsapp` was allowed, which made
    # attachments/quotes fail on group tickets with "Ticket nao e WhatsApp".
    if ticket.get("channel") not in ("whatsapp", "whatsapp_group"):
        raise HTTPException(status_code=400, detail="Ticket nao e WhatsApp")
    filename = (payload.get("filename") or "arquivo").strip()
    mimetype = (payload.get("mimetype") or "application/octet-stream").strip()
    data_b64 = payload.get("data_base64") or ""
    caption = (payload.get("caption") or "").strip()
    if not data_b64:
        raise HTTPException(status_code=400, detail="data_base64 obrigatorio")
    is_audio = mimetype.startswith("audio/")
    is_image = mimetype.startswith("image/")
    attachment_kind = "audio" if is_audio else ("image" if is_image else (
        "video" if mimetype.startswith("video/") else "document"
    ))
    # Persist the bytes to object storage so the chat can render an inline
    # player/preview via a normal URL (no base64 inflation in the ticket
    # JSON, no auth wall on <audio src>).
    from routes.channels_routes import _persist_inbound_media
    saved = await _persist_inbound_media(
        db, user["company_id"], data_b64,
        mimetype=mimetype, kind=attachment_kind, filename=filename,
    )
    message_id = str(uuid.uuid4())
    message = {
        "id": message_id,
        "content": caption or ("[Audio]" if is_audio else f"[{attachment_kind}] {filename}"),
        "sender_type": "agent",
        "sender_id": user["id"],
        "sender_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "delivery_status": "pending",
        "type": attachment_kind,
        "attachment_kind": attachment_kind,
        "attachment_filename": filename,
        "attachment_mimetype": mimetype,
    }
    # Surface as `media_*` so the chat UI renders the playable bubble
    # (the same fields used for inbound media coming from the webhook).
    if saved:
        message["media_url"] = saved["url"]
        message["media_kind"] = attachment_kind
        message["media_mimetype"] = saved["mimetype"]
        message["media_filename"] = saved["filename"]
        message["media_size"] = saved["size"]
    await db.tickets.update_one(
        {"id": ticket_id},
        {"$push": {"messages": message}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    # Sending any media counts as the operator taking over — pause the bot
    # if the company opted in. Same behavior as the text-message endpoint.
    try:
        from bot_pause import pause_bot_on_ticket_if_enabled
        await pause_bot_on_ticket_if_enabled(
            db, ticket, reason="agent_media_platform"
        )
    except Exception as e:
        import logging as _lg
        _lg.getLogger(__name__).warning(f"[bot_pause] media-send failed: {e}")
    # Try forwarding to WA microservice
    conn_id = ticket.get("connection_id")
    if not conn_id:
        return {"queued": True, "message": message}
    wa_url = os.environ.get("WA_SERVICE_URL", "http://localhost:3001")
    target_phone = ticket.get("customer_phone")
    try:
        import httpx
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                f"{wa_url}/instances/{conn_id}/send-media",
                json={
                    "phone": target_phone,
                    "filename": filename,
                    "mimetype": mimetype,
                    "data_base64": data_b64,
                    "caption": caption or "",
                },
            )
            if resp.status_code == 200 and resp.json().get("success"):
                msg_id_wa = resp.json().get("message_id")
                await db.tickets.update_one(
                    {"id": ticket_id, "messages.id": message_id},
                    {"$set": {
                        "messages.$.delivery_status": "sent",
                        "messages.$.wa_message_id": msg_id_wa,
                    }},
                )
                message["delivery_status"] = "sent"
                message["wa_message_id"] = msg_id_wa
            else:
                err = (resp.json() or {}).get("error") or f"HTTP {resp.status_code}"
                await db.tickets.update_one(
                    {"id": ticket_id, "messages.id": message_id},
                    {"$set": {"messages.$.delivery_status": "failed", "messages.$.delivery_error": err}},
                )
                message["delivery_status"] = "failed"
                message["delivery_error"] = err
    except Exception as e:
        await db.tickets.update_one(
            {"id": ticket_id, "messages.id": message_id},
            {"$set": {"messages.$.delivery_status": "failed", "messages.$.delivery_error": str(e)}},
        )
        message["delivery_status"] = "failed"
        message["delivery_error"] = str(e)
    # Don't return the base64 in the response to avoid 50MB JSON
    response = {**message}
    response.pop("attachment_data_b64", None)
    return response


@router.post("/tickets/{ticket_id}/tags/add")
async def add_tag_to_ticket(
    ticket_id: str,
    data: TicketTagToggle,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    res = await db.tickets.update_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"$addToSet": {"tags": data.tag}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")
    return await db.tickets.find_one({"id": ticket_id}, {"_id": 0})


@router.post("/tickets/{ticket_id}/tags/remove")
async def remove_tag_from_ticket(
    ticket_id: str,
    data: TicketTagToggle,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    res = await db.tickets.update_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"$pull": {"tags": data.tag}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")
    return await db.tickets.find_one({"id": ticket_id}, {"_id": 0})

# Kanban
@router.get("/kanban")
async def get_kanban(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    query = {"company_id": user["company_id"]}
    vis = _ticket_visibility_filter(user)
    if vis:
        query.update(vis)
    tickets = await db.tickets.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    kanban = {
        TicketStatus.EM_COBRANCA: [],
        TicketStatus.PAGO: [],
        TicketStatus.BLOQUEADO: [],
        TicketStatus.PROPOSTA: [],
        TicketStatus.ABERTO: [],
        TicketStatus.FECHADO: []
    }
    
    for ticket in tickets:
        status = ticket.get("status", TicketStatus.ABERTO)
        if status in kanban:
            kanban[status].append(ticket)
    
    return kanban

# AI Agent
@router.post("/ai/chat", response_model=AIChatResponse)
async def ai_chat(
    data: AIChatRequest,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    # Get ticket context
    ticket = await db.tickets.find_one({"id": data.ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket não encontrado")
    
    # Get or create AI conversation session
    session_id = data.session_id or str(uuid.uuid4())
    
    # Get conversation history
    conversation = await db.ai_conversations.find_one(
        {"ticket_id": data.ticket_id, "session_id": session_id}
    )
    
    if not conversation:
        conversation = {
            "id": str(uuid.uuid4()),
            "company_id": user["company_id"],
            "ticket_id": data.ticket_id,
            "session_id": session_id,
            "messages": [],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.ai_conversations.insert_one(conversation)
    
    # Prepare AI context
    system_message = f"""Você é um assistente de atendimento ao cliente.
    
Contexto do Ticket:
- Cliente: {ticket['customer_name']}
- Telefone: {ticket['customer_phone']}
- Status: {ticket['status']}
- Descrição: {ticket.get('description', 'N/A')}

Sua função é ajudar o atendente a responder o cliente de forma profissional e útil."""
    
    # Use Emergent LLM Key
    emergent_key = os.environ.get("EMERGENT_LLM_KEY")
    if not emergent_key:
        raise HTTPException(status_code=500, detail="Chave de API não configurada")
    
    try:
        chat = LlmChat(
            api_key=emergent_key,
            session_id=session_id,
            system_message=system_message
        ).with_model("openai", "gpt-5.2")
        
        user_message = UserMessage(text=data.message)
        response = await chat.send_message(user_message)
        
        # Save to conversation history
        new_messages = [
            {
                "role": "user",
                "content": data.message,
                "timestamp": datetime.now(timezone.utc).isoformat()
            },
            {
                "role": "assistant",
                "content": response,
                "timestamp": datetime.now(timezone.utc).isoformat()
            }
        ]
        
        await db.ai_conversations.update_one(
            {"session_id": session_id},
            {"$push": {"messages": {"$each": new_messages}}}
        )
        
        return AIChatResponse(response=response, session_id=session_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar chat: {str(e)}")

# Quick Responses
@router.get("/quick-responses")
async def list_quick_responses(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    responses = await db.quick_responses.find(
        {"company_id": user["company_id"]},
        {"_id": 0}
    ).to_list(1000)
    return responses

@router.post("/quick-responses")
async def create_quick_response(
    data: QuickResponseCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    response = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "title": data.title,
        "content": data.content,
        "shortcut": data.shortcut,
        "attachment_filename": data.attachment_filename or None,
        "attachment_mimetype": data.attachment_mimetype or None,
        "attachment_data_b64": data.attachment_data_b64 or None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.quick_responses.insert_one(response)
    return {k: v for k, v in response.items() if k != "_id"}


# 2026-02-28 — Edit/Delete de respostas rapidas. Antes so existia
# GET + POST; operador nao conseguia corrigir texto/atalho. Substitui
# anexo se o usuario subir um novo (comportamento intuitivo).
class QuickResponseUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    shortcut: Optional[str] = None
    attachment_filename: Optional[str] = None
    attachment_mimetype: Optional[str] = None
    attachment_data_b64: Optional[str] = None


@router.put("/quick-responses/{response_id}")
async def update_quick_response(
    response_id: str,
    data: QuickResponseUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    payload = data.model_dump(exclude_unset=True)
    update_set: dict = {}
    for field in ("title", "content", "shortcut"):
        if field in payload:
            update_set[field] = payload[field]
    # Anexo: se vier `attachment_data_b64` nao vazio, substitui (junto com
    # nome+mimetype); se vier explicitamente vazio, REMOVE.
    if "attachment_data_b64" in payload:
        b64 = payload.get("attachment_data_b64") or ""
        if b64:
            update_set["attachment_filename"] = payload.get("attachment_filename") or ""
            update_set["attachment_mimetype"] = payload.get("attachment_mimetype") or ""
            update_set["attachment_data_b64"] = b64
        else:
            update_set["attachment_filename"] = None
            update_set["attachment_mimetype"] = None
            update_set["attachment_data_b64"] = None
    if not update_set:
        raise HTTPException(status_code=400, detail="Nada para atualizar")
    update_set["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.quick_responses.update_one(
        {"id": response_id, "company_id": user["company_id"]},
        {"$set": update_set},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Resposta nao encontrada")
    doc = await db.quick_responses.find_one(
        {"id": response_id, "company_id": user["company_id"]}, {"_id": 0}
    )
    return doc


@router.delete("/quick-responses/{response_id}")
async def delete_quick_response(
    response_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    result = await db.quick_responses.delete_one(
        {"id": response_id, "company_id": user["company_id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Resposta nao encontrada")
    return {"deleted": response_id}


# === CAMPAIGN GLOBAL SETTINGS (anti-block policy per company) ===
class CampaignSettingsUpdate(BaseModel):
    anti_block: dict


@router.get("/campaign-settings")
async def get_campaign_settings(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    doc = await db.campaign_settings.find_one({"company_id": user["company_id"]}, {"_id": 0})
    if not doc:
        doc = {
            "company_id": user["company_id"],
            "anti_block": {
                "enabled": True,
                "interval_min_seconds": 30,
                "interval_max_seconds": 90,
                "burst_size": 50,
                "burst_pause_seconds": 300,
                "daily_limit": 250,
                "hourly_limit": 50,
                "escalate_after": 100,
                "escalate_factor": 1.5,
                "only_with_phone_validated": True,
            }
        }
    return doc


@router.put("/campaign-settings")
async def update_campaign_settings(
    data: CampaignSettingsUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    await db.campaign_settings.update_one(
        {"company_id": user["company_id"]},
        {"$set": {"anti_block": data.anti_block,
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    doc = await db.campaign_settings.find_one({"company_id": user["company_id"]}, {"_id": 0})
    return doc


# === BOT-PAUSE SETTINGS (per company) ===
# Controls whether the WhatsApp bot (Flowbuilder runtime) automatically
# steps aside when an operator sends a message — either via the CRM UI or
# via their linked phone. See /app/backend/bot_pause.py for the runtime
# integration. The toggle defaults to ON for new tenants; existing tenants
# inherit the same default until they explicitly change it via this endpoint.
class BotSettingsUpdate(BaseModel):
    pause_bot_on_human_intervention: bool


@router.get("/company/bot-settings")
async def get_company_bot_settings(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    from bot_pause import is_pause_setting_enabled
    enabled = await is_pause_setting_enabled(db, user["company_id"])
    return {"pause_bot_on_human_intervention": enabled}


@router.put("/company/bot-settings")
async def update_company_bot_settings(
    data: BotSettingsUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    # Only company admins (owner/admin) can toggle this. Regular operators
    # would see the toggle disabled in the UI, but enforce server-side too.
    role = (user.get("role") or "").lower()
    if role not in ("company_admin", "owner", "super_admin", "admin"):
        raise HTTPException(403, "Apenas administradores podem alterar esta configuracao")
    await db.companies.update_one(
        {"id": user["company_id"]},
        {"$set": {
            "pause_bot_on_human_intervention": bool(data.pause_bot_on_human_intervention),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    return {"pause_bot_on_human_intervention": bool(data.pause_bot_on_human_intervention)}


# === TICKET LIFECYCLE SETTINGS (per company) ===
# Controls SGP-gateway auto-close (every successful outbound via the public
# `/api/sgp/gateway/send/{token}` endpoint will fechar the ticket right
# away) and an inactivity timeout (after N hours without any new message,
# the scheduler closes the ticket automatically). Both default to safe
# values: auto_close=False, timeout=0 (disabled). New customers can flip
# them ON from /configuracoes.
class TicketLifecycleSettingsUpdate(BaseModel):
    sgp_gateway_auto_close: Optional[bool] = None
    ticket_auto_close_hours: Optional[int] = None
    ticket_auto_close_message: Optional[str] = None
    # 2026-05-28 — Quando True, fechamento MANUAL tambem envia a mensagem.
    send_close_message_on_manual: Optional[bool] = None


@router.get("/company/ticket-settings")
async def get_company_ticket_settings(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    comp = await db.companies.find_one(
        {"id": user["company_id"]},
        {"_id": 0, "sgp_gateway_auto_close": 1, "ticket_auto_close_hours": 1,
         "ticket_auto_close_message": 1, "send_close_message_on_manual": 1},
    ) or {}
    return {
        "sgp_gateway_auto_close": bool(comp.get("sgp_gateway_auto_close", False)),
        "ticket_auto_close_hours": int(comp.get("ticket_auto_close_hours") or 0),
        "ticket_auto_close_message": comp.get("ticket_auto_close_message") or "",
        "send_close_message_on_manual": bool(comp.get("send_close_message_on_manual", False)),
    }


@router.put("/company/ticket-settings")
async def update_company_ticket_settings(
    data: TicketLifecycleSettingsUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    role = (user.get("role") or "").lower()
    if role not in ("company_admin", "owner", "super_admin", "admin"):
        raise HTTPException(403, "Apenas administradores podem alterar esta configuracao")
    update: dict = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if data.sgp_gateway_auto_close is not None:
        update["sgp_gateway_auto_close"] = bool(data.sgp_gateway_auto_close)
    if data.ticket_auto_close_hours is not None:
        hours = int(data.ticket_auto_close_hours)
        if hours < 0 or hours > 24 * 30:  # cap at 30 days
            raise HTTPException(400, "ticket_auto_close_hours fora do intervalo permitido (0-720)")
        update["ticket_auto_close_hours"] = hours
    if data.ticket_auto_close_message is not None:
        # Trim e cap em 1000 chars pra evitar payload abusivo.
        update["ticket_auto_close_message"] = (data.ticket_auto_close_message or "")[:1000]
    if data.send_close_message_on_manual is not None:
        update["send_close_message_on_manual"] = bool(data.send_close_message_on_manual)
    await db.companies.update_one(
        {"id": user["company_id"]},
        {"$set": update},
    )
    comp = await db.companies.find_one(
        {"id": user["company_id"]},
        {"_id": 0, "sgp_gateway_auto_close": 1, "ticket_auto_close_hours": 1,
         "ticket_auto_close_message": 1, "send_close_message_on_manual": 1},
    ) or {}
    return {
        "sgp_gateway_auto_close": bool(comp.get("sgp_gateway_auto_close", False)),
        "ticket_auto_close_hours": int(comp.get("ticket_auto_close_hours") or 0),
        "ticket_auto_close_message": comp.get("ticket_auto_close_message") or "",
        "send_close_message_on_manual": bool(comp.get("send_close_message_on_manual", False)),
    }


# === BOT-PAUSE per-ticket overrides ===
# Operator can manually resume the bot on a paused ticket (e.g. they were
# investigating an issue and want the flow to take over again) or pause it
# without sending a message (rare, but useful for VIP customers).
@router.post("/tickets/{ticket_id}/bot-pause")
async def toggle_ticket_bot_pause(
    ticket_id: str,
    payload: dict,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Body: {"paused": true|false}. When `paused=false`, clears the pause
    flags. When `paused=true`, sets bot_paused on the ticket (and clears
    active_flow_node_id like the automatic pauser does)."""
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]}, {"_id": 0})
    if not ticket:
        raise HTTPException(404, "Ticket nao encontrado")
    target = bool(payload.get("paused"))
    now = datetime.now(timezone.utc).isoformat()
    if target:
        await db.tickets.update_one(
            {"id": ticket_id},
            {"$set": {
                "bot_paused": True,
                "bot_paused_at": now,
                "bot_paused_reason": "manual_toggle",
                "active_flow_node_id": None,
                "updated_at": now,
            }},
        )
    else:
        from bot_pause import resume_bot_on_ticket
        await resume_bot_on_ticket(db, ticket_id)
    updated = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    return {"bot_paused": bool(updated.get("bot_paused")),
            "bot_paused_at": updated.get("bot_paused_at"),
            "bot_paused_reason": updated.get("bot_paused_reason")}


# Campaigns
@router.get("/campaigns")
async def list_campaigns(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    campaigns = await db.campaigns.find(
        {"company_id": user["company_id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    # Enrich with connection name + list name for the list view
    conn_ids = list({c.get("connection_id") for c in campaigns if c.get("connection_id")})
    list_ids = list({c.get("contact_list_id") for c in campaigns if c.get("contact_list_id")})
    conns = {}
    if conn_ids:
        async for cn in db.channel_connections.find({"id": {"$in": conn_ids}}, {"_id":0,"id":1,"name":1}):
            conns[cn["id"]] = cn.get("name")
    lists = {}
    if list_ids:
        async for ln in db.contact_lists.find({"id": {"$in": list_ids}}, {"_id":0,"id":1,"name":1}):
            lists[ln["id"]] = ln.get("name")
    for c in campaigns:
        c["connection_name"] = conns.get(c.get("connection_id"), "-")
        c["contact_list_name"] = lists.get(c.get("contact_list_id"), "-")
    return campaigns


@router.post("/campaigns")
async def create_campaign(
    data: CampaignCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    campaign = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "name": data.name,
        "type": data.type,
        "audience_mode": data.audience_mode,
        "tag_ids": data.tag_ids or [],
        "contact_list_id": data.contact_list_id,
        "connection_id": data.connection_id,
        "scheduled_at": data.scheduled_at,
        "confirmation_enabled": data.confirmation_enabled,
        "open_ticket": data.open_ticket,
        "assigned_user_id": data.assigned_user_id,
        "queue_id": data.queue_id,
        "ticket_status": data.ticket_status or "fechado",
        "messages": data.messages or ([data.message_template] if data.message_template else []),
        "attachment_url": data.attachment_url,
        "attachment_filename": data.attachment_filename,
        "anti_block": (data.anti_block.model_dump() if data.anti_block else {
            "enabled": True, "interval_min_seconds": 30, "interval_max_seconds": 90,
            "burst_size": 50, "burst_pause_seconds": 300, "daily_limit": 250,
            "hourly_limit": 50, "escalate_after": 100, "escalate_factor": 1.5,
            "only_with_phone_validated": True,
        }),
        # 2026-02-28 — Modo Disparo em Massa armazenado na campanha.
        "bulk_config": (data.bulk_config.model_dump() if data.bulk_config else None),
        "status": "programada" if data.scheduled_at else "draft",
        "sent_count": 0,
        "failed_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.campaigns.insert_one(campaign)
    return {k: v for k, v in campaign.items() if k != "_id"}


@router.put("/campaigns/{campaign_id}")
async def update_campaign(
    campaign_id: str,
    data: CampaignUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    update = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="Sem dados")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.campaigns.update_one(
        {"id": campaign_id, "company_id": user["company_id"]}, {"$set": update}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Campanha nao encontrada")
    return await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})


@router.delete("/campaigns/{campaign_id}")
async def delete_campaign(
    campaign_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    res = await db.campaigns.delete_one({"id": campaign_id, "company_id": user["company_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Campanha nao encontrada")
    return {"message": "Removida"}


async def _resolve_campaign_audience(db: AsyncIOMotorDatabase, company_id: str, c: dict) -> List[dict]:
    """Return [{name, phone}] based on campaign filters."""
    mode = c.get("audience_mode") or "tags"
    out: List[dict] = []
    if mode == "list":
        cl_id = c.get("contact_list_id")
        if not cl_id:
            return []
        cl = await db.contact_lists.find_one({"id": cl_id, "company_id": company_id}, {"_id": 0})
        if not cl:
            return []
        for it in (cl.get("contacts") or []):
            if it.get("phone"):
                out.append({"name": it.get("name") or "", "phone": it["phone"]})
        return out

    # Pull all clients for the company once
    clients = await db.clients.find({"company_id": company_id}, {"_id": 0}).to_list(10000)

    if mode == "all":
        for cli in clients:
            if cli.get("phone"):
                out.append({"name": cli.get("name") or "", "phone": cli["phone"]})
        return out

    if mode == "no_tag":
        for cli in clients:
            if cli.get("phone") and not (cli.get("tags") or []):
                out.append({"name": cli.get("name") or "", "phone": cli["phone"]})
        return out

    # mode == 'tags'
    tag_ids = c.get("tag_ids") or []
    if not tag_ids:
        return []
    # Resolve tag names
    tag_docs = await db.tags.find({"id": {"$in": tag_ids}, "company_id": company_id}, {"_id":0,"name":1}).to_list(50)
    tag_names = {t["name"] for t in tag_docs}
    for cli in clients:
        client_tags = set(cli.get("tags") or [])
        if cli.get("phone") and (client_tags & tag_names):
            out.append({"name": cli.get("name") or "", "phone": cli["phone"]})
    # Also include tickets with such tags (they may not yet be "clients")
    async for t in db.tickets.find(
        {"company_id": company_id, "tags": {"$in": list(tag_names)}},
        {"_id": 0, "customer_name": 1, "customer_phone": 1}
    ):
        if t.get("customer_phone"):
            out.append({"name": t.get("customer_name") or "", "phone": t["customer_phone"]})
    # Dedup by phone
    seen = set()
    dedup = []
    for it in out:
        if it["phone"] in seen:
            continue
        seen.add(it["phone"])
        dedup.append(it)
    return dedup


@router.post("/campaigns/{campaign_id}/preview-audience")
async def preview_campaign_audience(
    campaign_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    camp = await db.campaigns.find_one({"id": campaign_id, "company_id": user["company_id"]}, {"_id": 0})
    if not camp:
        raise HTTPException(status_code=404, detail="Campanha nao encontrada")
    audience = await _resolve_campaign_audience(db, user["company_id"], camp)
    return {"count": len(audience), "preview": audience[:50]}


@router.post("/campaigns/{campaign_id}/run")
async def run_campaign_now(
    campaign_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    """Send the campaign immediately (ignores scheduled_at).

    2026-02-28 — Se a campanha tem `bulk_config.enabled=true`, encaminha
    para o bulk dispatcher (fila persistente, multi-conexao, spintax,
    janela, opt-out, daily cap). Caso contrario, usa o pipeline classico.

    2026-06-25 — Modo classico agora SEMPRE roda async + grava cada
    destinatario em `campaign_deliveries` (status pending/sending/sent/
    failed). Isso permite (a) pause/resume, (b) visualizacao em tempo
    real do progresso pelo modal do "olhinho", (c) o scheduler agendar
    campanhas com `scheduled_at`.
    """
    camp = await db.campaigns.find_one({"id": campaign_id, "company_id": user["company_id"]}, {"_id": 0})
    if not camp:
        raise HTTPException(status_code=404, detail="Campanha nao encontrada")

    # ─── Bulk mode: roteia pro dispatcher de massa ──────────────────
    bulk_cfg = camp.get("bulk_config") or {}
    if bulk_cfg.get("enabled"):
        connection_ids = bulk_cfg.get("connection_ids") or ([camp.get("connection_id")] if camp.get("connection_id") else [])
        if not connection_ids:
            raise HTTPException(status_code=400, detail="Modo Disparo em Massa: selecione ao menos 1 conexao em bulk_config.connection_ids")
        from routes.bulk_routes import (
            create_bulk_job_from_campaign,
            JobFromCampaign,
            WindowConfig,
        )
        payload = JobFromCampaign(
            connection_ids=connection_ids,
            interval_min_sec=int(bulk_cfg.get("interval_min_sec", 8)),
            interval_max_sec=int(bulk_cfg.get("interval_max_sec", 25)),
            daily_cap_per_connection=int(bulk_cfg.get("daily_cap_per_connection", 800)),
            opt_out_keywords=bulk_cfg.get("opt_out_keywords") or ["PARAR", "SAIR", "DESCADASTRAR"],
            window=WindowConfig(
                enabled=bool(bulk_cfg.get("window_enabled", True)),
                start=bulk_cfg.get("window_start", "09:00"),
                end=bulk_cfg.get("window_end", "18:00"),
                days_of_week=bulk_cfg.get("window_days") or [0, 1, 2, 3, 4, 5],
            ),
            auto_start=True,
        )
        job = await create_bulk_job_from_campaign(campaign_id, payload, user=user, db=db)
        await db.campaigns.update_one(
            {"id": campaign_id}, {"$set": {"status": "em_execucao", "last_bulk_job_id": job.get("id")}}
        )
        return {
            "mode": "bulk",
            "job_id": job.get("id"),
            "audience": job.get("audience_size"),
            "message": "Disparo em Massa iniciado. Acompanhe na aba 'Disparos em Massa'.",
        }

    # ─── Classic mode: seed deliveries + run async ──────────────────
    result = await _fire_campaign_classic(db, camp)
    if result.get("error"):
        raise HTTPException(status_code=400, detail=result["error"])
    return result


async def _fire_campaign_classic(db: AsyncIOMotorDatabase, camp: dict) -> dict:
    """Seed deliveries + start async runner. Used by both /run and the
    scheduler. Returns {queued, total} or {error}.
    """
    company_id = camp["company_id"]
    campaign_id = camp["id"]
    audience = await _resolve_campaign_audience(db, company_id, camp)
    if not audience:
        return {"error": "Audiencia vazia"}
    msgs = [m for m in (camp.get("messages") or []) if m and m.strip()]
    # 2026-06-27 — Permite campanha "midia-only" (PDF/imagem sem caption).
    # Antes exigiamos pelo menos 1 mensagem; agora basta ter o anexo.
    if not msgs and not camp.get("attachment_url"):
        return {"error": "Sem mensagens nem anexo definidos"}

    conn_id = camp.get("connection_id")
    if not conn_id:
        c2 = await db.channel_connections.find_one(
            {"company_id": company_id, "type": "whatsapp", "status": "connected"}, {"_id": 0, "id": 1}
        )
        if not c2:
            return {"error": "Nenhuma conexao WhatsApp ativa"}
        conn_id = c2["id"]

    # Anti-block (campaign-level override OR company settings)
    ab = camp.get("anti_block") or {}
    if not ab:
        settings = await db.campaign_settings.find_one({"company_id": company_id}, {"_id": 0})
        ab = (settings or {}).get("anti_block") or {}
    daily_limit = max(1, int(ab.get("daily_limit", 250) or 250))
    if len(audience) > daily_limit:
        audience = audience[:daily_limit]

    # Wipe any prior deliveries for re-runs (e.g. operator hit Send again
    # after a failed campaign). The progress modal always shows the most
    # recent execution.
    await db.campaign_deliveries.delete_many({"campaign_id": campaign_id})
    now_iso = datetime.now(timezone.utc).isoformat()
    deliveries = [
        {
            "id": str(uuid.uuid4()),
            "campaign_id": campaign_id,
            "company_id": company_id,
            "name": p.get("name") or "",
            "phone": p["phone"],
            "status": "pending",
            "created_at": now_iso,
        }
        for p in audience
    ]
    if deliveries:
        await db.campaign_deliveries.insert_many(deliveries)

    await db.campaigns.update_one(
        {"id": campaign_id},
        {"$set": {
            "status": "em_execucao",
            "started_at": now_iso,
            "sent_count": 0,
            "failed_count": 0,
            "total_count": len(audience),
            "connection_id": conn_id,
        }, "$unset": {"completed_at": "", "error": ""}}
    )

    import asyncio as _asyncio
    _asyncio.create_task(_classic_runner(campaign_id, conn_id, msgs, ab, camp.get("attachment_url"), camp.get("attachment_filename")))
    return {"queued": True, "total": len(audience), "mode": "classic"}


async def _classic_runner(campaign_id: str, conn_id: str, msgs: list, ab: dict, attachment_url: str = None, attachment_filename: str = None):
    """Background dispatch loop with pause/cancel awareness.

    2026-06-27 — Now supports an optional attachment (image/PDF/doc). When
    `attachment_url` is present, we fetch the bytes ONCE and reuse them
    across all deliveries; the first message in `msgs` is sent as the
    media caption (if any), then remaining messages are sent as text.
    Empty messages array + attachment => media-only send.
    """
    import asyncio as _asyncio
    import random as _random
    import httpx as _httpx
    import os as _os
    import base64 as _b64
    from motor.motor_asyncio import AsyncIOMotorClient as _Cli
    from notifications import render_template as _render
    from wa_humanize import humanize_kwargs as _hum

    wa_url = _os.environ.get("WA_SERVICE_URL", "http://localhost:3002")
    cli = _Cli(_os.environ["MONGO_URL"])
    bdb = cli[_os.environ["DB_NAME"]]

    # Pre-load the attachment ONCE (instead of per recipient) — same bytes,
    # same mime, sent to many contacts. Saves a few hundred ms per send and
    # avoids hammering the object storage / disk.
    media_b64 = None
    media_mime = None
    if attachment_url:
        try:
            from routes.upload_routes import get_object as _get_object
            # attachment_url is "/api/upload/files/<path>" — strip the prefix
            # to get the storage path. Anything else (external URL): fetch via HTTP.
            if attachment_url.startswith("/api/upload/files/"):
                storage_path = attachment_url[len("/api/upload/files/"):]
                data, content_type = _get_object(storage_path)
                media_b64 = _b64.b64encode(data).decode("ascii")
                media_mime = content_type or "application/octet-stream"
            else:
                async with _httpx.AsyncClient(timeout=30.0) as _c:
                    rr = await _c.get(attachment_url)
                    if rr.status_code == 200:
                        media_b64 = _b64.b64encode(rr.content).decode("ascii")
                        media_mime = rr.headers.get("content-type") or "application/octet-stream"
            print(f"[classic_runner] campaign={campaign_id} attachment loaded ({len(media_b64 or '')} chars b64, mime={media_mime})")
        except Exception as e:
            print(f"[classic_runner] failed to load attachment for campaign={campaign_id}: {e}")
            media_b64 = None

    ab_enabled = ab.get("enabled", True)
    interval_min = max(0, int(ab.get("interval_min_seconds", 30) or 0))
    interval_max = max(interval_min, int(ab.get("interval_max_seconds", 90) or 0))
    burst_size = max(1, int(ab.get("burst_size", 50) or 1))
    burst_pause = max(0, int(ab.get("burst_pause_seconds", 300) or 0))
    escalate_after = max(0, int(ab.get("escalate_after", 100) or 0))
    escalate_factor = float(ab.get("escalate_factor", 1.5) or 1.0)

    sent_x, failed_x, count = 0, 0, 0

    async def _wait_if_paused():
        """If campaign is `pausada`, sleep until it's resumed or
        cancelled. Returns True if cancelled/finished (caller should
        exit), False otherwise."""
        while True:
            doc = await bdb.campaigns.find_one({"id": campaign_id}, {"_id": 0, "status": 1})
            st = (doc or {}).get("status")
            if st in ("cancelada", "concluida"):
                return True
            if st == "pausada":
                await _asyncio.sleep(5)
                continue
            return False

    try:
        async with _httpx.AsyncClient(timeout=30.0) as client:
            cursor = bdb.campaign_deliveries.find(
                {"campaign_id": campaign_id, "status": "pending"}, {"_id": 0}
            ).sort("created_at", 1)
            deliveries = await cursor.to_list(100000)
            total = len(deliveries)
            for d in deliveries:
                if await _wait_if_paused():
                    break
                await bdb.campaign_deliveries.update_one(
                    {"id": d["id"]}, {"$set": {"status": "sending"}}
                )
                ok_any = False
                last_err = ""
                # 2026-06-27 — If we have media, send it FIRST. The first
                # template message (if any) becomes the caption; the rest
                # are sent as separate text messages below.
                msgs_remaining = list(msgs)
                if media_b64:
                    caption_tpl = msgs_remaining.pop(0) if msgs_remaining else ""
                    caption_txt = _render(caption_tpl or "", {"nome": d.get("name") or "", "numero": d["phone"], "telefone": d["phone"]}) if caption_tpl else ""
                    try:
                        rr = await client.post(
                            f"{wa_url}/instances/{conn_id}/send-media",
                            json={
                                "phone": d["phone"],
                                "filename": attachment_filename or "anexo.bin",
                                "mimetype": media_mime,
                                "data_base64": media_b64,
                                "caption": caption_txt,
                            },
                            timeout=60.0,
                        )
                        rs = rr.json() if rr.status_code == 200 else {}
                        if rs.get("success"):
                            ok_any = True
                        else:
                            last_err = (rs.get("error") or f"media http {rr.status_code}")[:200]
                    except Exception as e:
                        last_err = f"media: {str(e)[:200]}"
                for tpl in msgs_remaining:
                    mtxt = _render(tpl or "", {"nome": d.get("name") or "", "numero": d["phone"], "telefone": d["phone"]})
                    try:
                        hum = await _hum(bdb, conn_id)
                        rr = await client.post(
                            f"{wa_url}/instances/{conn_id}/send",
                            json={"phone": d["phone"], "message": mtxt, **hum},
                        )
                        rs = rr.json() if rr.status_code == 200 else {}
                        if rs.get("success"):
                            ok_any = True
                        else:
                            last_err = (rs.get("error") or f"http {rr.status_code}")[:200]
                    except Exception as e:
                        last_err = str(e)[:200]
                if ok_any:
                    sent_x += 1
                    await bdb.campaign_deliveries.update_one(
                        {"id": d["id"]},
                        {"$set": {"status": "sent", "sent_at": datetime.now(timezone.utc).isoformat()}}
                    )
                else:
                    failed_x += 1
                    await bdb.campaign_deliveries.update_one(
                        {"id": d["id"]},
                        {"$set": {"status": "failed", "error": last_err or "unknown",
                                  "failed_at": datetime.now(timezone.utc).isoformat()}}
                    )
                await bdb.campaigns.update_one(
                    {"id": campaign_id},
                    {"$set": {"sent_count": sent_x, "failed_count": failed_x}}
                )
                count += 1
                if not ab_enabled or count >= total:
                    continue
                # Pause check before sleeping so resume doesn't have to wait
                if await _wait_if_paused():
                    break
                cur_min, cur_max = interval_min, interval_max
                if escalate_after and count > escalate_after:
                    cur_min = int(cur_min * escalate_factor)
                    cur_max = int(cur_max * escalate_factor)
                if burst_size and count % burst_size == 0:
                    await _asyncio.sleep(burst_pause)
                else:
                    await _asyncio.sleep(_random.randint(cur_min, max(cur_min, cur_max)))

        # Only mark concluida if not cancelled
        doc = await bdb.campaigns.find_one({"id": campaign_id}, {"_id": 0, "status": 1})
        if (doc or {}).get("status") not in ("cancelada",):
            await bdb.campaigns.update_one(
                {"id": campaign_id},
                {"$set": {"status": "concluida", "sent_count": sent_x, "failed_count": failed_x,
                          "completed_at": datetime.now(timezone.utc).isoformat()}}
            )
    except Exception as e:
        await bdb.campaigns.update_one(
            {"id": campaign_id},
            {"$set": {"status": "cancelada", "error": str(e)[:200],
                      "sent_count": sent_x, "failed_count": failed_x}}
        )


@router.post("/campaigns/{campaign_id}/pause")
async def pause_campaign(
    campaign_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    res = await db.campaigns.update_one(
        {"id": campaign_id, "company_id": user["company_id"], "status": "em_execucao"},
        {"$set": {"status": "pausada", "paused_at": datetime.now(timezone.utc).isoformat()}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=400, detail="Campanha nao esta em execucao")
    return {"message": "Campanha pausada"}


@router.post("/campaigns/{campaign_id}/resume")
async def resume_campaign(
    campaign_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    res = await db.campaigns.update_one(
        {"id": campaign_id, "company_id": user["company_id"], "status": "pausada"},
        {"$set": {"status": "em_execucao"}, "$unset": {"paused_at": ""}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=400, detail="Campanha nao esta pausada")
    return {"message": "Campanha retomada"}


@router.post("/campaigns/{campaign_id}/cancel")
async def cancel_campaign(
    campaign_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    res = await db.campaigns.update_one(
        {"id": campaign_id, "company_id": user["company_id"], "status": {"$in": ["em_execucao", "pausada", "programada"]}},
        {"$set": {"status": "cancelada", "cancelled_at": datetime.now(timezone.utc).isoformat()}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=400, detail="Campanha nao pode ser cancelada")
    return {"message": "Campanha cancelada"}


@router.get("/campaigns/{campaign_id}/progress")
async def get_campaign_progress(
    campaign_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Real-time progress for the campaign — totals + recent deliveries
    per status. Used by the live "olhinho" modal."""
    camp = await db.campaigns.find_one(
        {"id": campaign_id, "company_id": user["company_id"]},
        {"_id": 0, "id": 1, "name": 1, "status": 1, "scheduled_at": 1, "started_at": 1,
         "completed_at": 1, "sent_count": 1, "failed_count": 1, "total_count": 1, "error": 1},
    )
    if not camp:
        raise HTTPException(status_code=404, detail="Campanha nao encontrada")
    pipeline = [
        {"$match": {"campaign_id": campaign_id, "company_id": user["company_id"]}},
        {"$group": {"_id": "$status", "n": {"$sum": 1}}},
    ]
    counts = {"pending": 0, "sending": 0, "sent": 0, "failed": 0}
    async for row in db.campaign_deliveries.aggregate(pipeline):
        counts[row["_id"]] = row["n"]
    total = sum(counts.values()) or int(camp.get("total_count") or 0)
    # Latest items per bucket (cap at 200 to keep payload small)
    async def _list(st, limit=200):
        return await db.campaign_deliveries.find(
            {"campaign_id": campaign_id, "company_id": user["company_id"], "status": st},
            {"_id": 0, "id": 1, "name": 1, "phone": 1, "status": 1, "error": 1, "sent_at": 1, "failed_at": 1},
        ).sort([("sent_at", -1), ("failed_at", -1), ("created_at", 1)]).limit(limit).to_list(limit)
    sent = await _list("sent")
    failed = await _list("failed")
    pending = await _list("pending", limit=50)
    sending = await _list("sending", limit=20)
    return {
        "campaign": camp,
        "totals": {**counts, "total": total},
        "sent": sent,
        "failed": failed,
        "pending": pending,
        "sending": sending,
    }


# === Excel import / template for contact lists ===
@router.get("/contact-lists/template.xlsx")
async def contact_list_template(user: dict = Depends(get_current_user)):
    """Returns an .xlsx template with columns: Nome, Telefone, Email."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatos"
    ws.append(["Nome", "Telefone", "Email"])
    ws.append(["Joao Silva", "5511999999999", "joao@email.com"])
    ws.append(["Maria Souza", "5511988888888", ""])
    for col, width in zip("ABC", (25, 22, 28)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo-contatos.xlsx"'},
    )


@router.post("/contact-lists/import-excel")
async def import_contact_list_excel(
    file: UploadFile = File(...),
    name: str = "Lista Importada",
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Parse an uploaded xlsx and create a new contact list.

    Accepts headers (case-insensitive): nome|name, telefone|phone, email.
    Telefone is required; rows without a phone are skipped. Duplicates
    (same phone) are de-duped, keeping the first non-empty name.
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Arquivo precisa ser .xlsx")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Planilha invalida: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia")
    headers = [str(h or "").strip().lower() for h in rows[0]]

    def col(name_options):
        for nm in name_options:
            if nm in headers:
                return headers.index(nm)
        return -1

    i_name = col(["nome", "name"])
    i_phone = col(["telefone", "phone", "celular", "whatsapp"])
    i_email = col(["email", "e-mail"])
    if i_phone < 0:
        raise HTTPException(status_code=400, detail="Coluna obrigatoria 'Telefone' nao encontrada")

    seen = set()
    contacts = []
    skipped = 0
    for r in rows[1:]:
        if not r:
            continue
        phone_raw = r[i_phone] if i_phone < len(r) else None
        if phone_raw is None or str(phone_raw).strip() == "":
            skipped += 1
            continue
        phone = re.sub(r"\D", "", str(phone_raw))
        if not phone:
            skipped += 1
            continue
        if phone in seen:
            continue
        seen.add(phone)
        nm = ""
        if i_name >= 0 and i_name < len(r) and r[i_name]:
            nm = str(r[i_name]).strip()
        em = ""
        if i_email >= 0 and i_email < len(r) and r[i_email]:
            em = str(r[i_email]).strip()
        contacts.append({"name": nm, "phone": phone, "email": em})

    doc = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "name": name or "Lista Importada",
        "description": f"Importada de {file.filename}",
        "contacts": contacts,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.contact_lists.insert_one(doc)
    return {
        "id": doc["id"],
        "name": doc["name"],
        "imported_count": len(contacts),
        "skipped_count": skipped,
    }





# === QUEUES (Filas & Chatbot) ===
class QueueCreate(BaseModel):
    name: str
    color: Optional[str] = "#4F46E5"
    description: Optional[str] = ""
    welcome_message: Optional[str] = ""
    bot_flow_id: Optional[str] = None
    connection_ids: Optional[List[str]] = None  # M3: queue → 1+ WhatsApp connections


class QueueUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    description: Optional[str] = None
    welcome_message: Optional[str] = None
    bot_flow_id: Optional[str] = None
    connection_ids: Optional[List[str]] = None


@router.get("/queues")
async def list_queues(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    return await db.queues.find({"company_id": user["company_id"]}, {"_id": 0}).sort("name", 1).to_list(200)


@router.post("/queues")
async def create_queue(
    data: QueueCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    doc = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "name": data.name,
        "color": data.color or "#4F46E5",
        "description": data.description or "",
        "welcome_message": data.welcome_message or "",
        "bot_flow_id": data.bot_flow_id,
        "connection_ids": data.connection_ids or [],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.queues.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@router.put("/queues/{queue_id}")
async def update_queue(
    queue_id: str,
    data: QueueUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    update = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="Sem dados")
    res = await db.queues.update_one(
        {"id": queue_id, "company_id": user["company_id"]}, {"$set": update}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fila nao encontrada")
    return await db.queues.find_one({"id": queue_id}, {"_id": 0})


@router.delete("/queues/{queue_id}")
async def delete_queue(
    queue_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    res = await db.queues.delete_one({"id": queue_id, "company_id": user["company_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fila nao encontrada")
    return {"message": "Removida"}


# === CONTACT LISTS ===
class ContactItem(BaseModel):
    name: Optional[str] = ""
    phone: str


class ContactListCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    contacts: Optional[List[ContactItem]] = None


class ContactListUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    contacts: Optional[List[ContactItem]] = None


@router.get("/contact-lists")
async def list_contact_lists(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    out = await db.contact_lists.find({"company_id": user["company_id"]}, {"_id": 0}).sort("name", 1).to_list(200)
    for o in out:
        o["count"] = len(o.get("contacts") or [])
    return out


@router.post("/contact-lists")
async def create_contact_list(
    data: ContactListCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    contacts = [c.model_dump() for c in (data.contacts or [])]
    doc = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "name": data.name,
        "description": data.description or "",
        "contacts": contacts,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.contact_lists.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@router.put("/contact-lists/{list_id}")
async def update_contact_list(
    list_id: str,
    data: ContactListUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    update = {}
    if data.name is not None: update["name"] = data.name
    if data.description is not None: update["description"] = data.description
    if data.contacts is not None: update["contacts"] = [c.model_dump() for c in data.contacts]
    if not update:
        raise HTTPException(status_code=400, detail="Sem dados")
    res = await db.contact_lists.update_one(
        {"id": list_id, "company_id": user["company_id"]}, {"$set": update}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lista nao encontrada")
    return await db.contact_lists.find_one({"id": list_id}, {"_id": 0})


@router.delete("/contact-lists/{list_id}")
async def delete_contact_list(
    list_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    res = await db.contact_lists.delete_one({"id": list_id, "company_id": user["company_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lista nao encontrada")
    return {"message": "Removida"}


# === RETRY MESSAGE ===
@router.post("/tickets/{ticket_id}/messages/{message_id}/retry")
async def retry_ticket_message(
    ticket_id: str,
    message_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    ticket = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")
    msg = next((m for m in (ticket.get("messages") or []) if m.get("id") == message_id), None)
    if not msg:
        raise HTTPException(status_code=404, detail="Mensagem nao encontrada")
    if msg.get("sender_type") != "agent" or ticket.get("channel") != "whatsapp":
        raise HTTPException(status_code=400, detail="Reenvio so para mensagens de agente em WhatsApp")
    if not ticket.get("customer_phone"):
        raise HTTPException(status_code=400, detail="Sem telefone")

    import httpx as _httpx
    import os as _os
    wa_url = _os.environ.get("WA_SERVICE_URL", "http://localhost:3002")
    # Prefer ticket's connection_id, but fall back to any currently
    # CONNECTED connection in the tenant when the original is stale,
    # disconnected, or deleted. The retry button in production was
    # silently failing because old tickets pointed to long-rotated
    # connection_ids — now we always pick a usable one if the operator
    # didn't explicitly switch via the ConnectionSwitcher.
    conn_id = ticket.get("connection_id")
    if conn_id:
        conn_doc = await db.channel_connections.find_one(
            {"id": conn_id, "company_id": user["company_id"]},
            {"_id": 0, "id": 1, "status": 1, "name": 1},
        )
        if not conn_doc or (conn_doc.get("status") or "").lower() != "connected":
            conn_id = None  # force fallback
    if not conn_id:
        c2 = await db.channel_connections.find_one(
            {"company_id": user["company_id"], "type": "whatsapp", "status": "connected"},
            {"_id": 0, "id": 1},
        )
        conn_id = c2["id"] if c2 else None
    if not conn_id:
        raise HTTPException(status_code=400, detail="Nenhuma conexao WhatsApp conectada no momento")
    try:
        async with _httpx.AsyncClient(timeout=30.0) as client:
            r = await client.post(f"{wa_url}/instances/{conn_id}/send",
                                  json={"phone": ticket["customer_phone"], "message": msg["content"]})
            res = r.json() if r.status_code == 200 else {}
        if res.get("success"):
            await db.tickets.update_one(
                {"id": ticket_id, "messages.id": message_id},
                {"$set": {"messages.$.delivery_status": "sent", "messages.$.delivery_error": None}}
            )
            return {"ok": True}
        else:
            err = res.get("error", "Falha ao reenviar")
            await db.tickets.update_one(
                {"id": ticket_id, "messages.id": message_id},
                {"$set": {"messages.$.delivery_status": "failed", "messages.$.delivery_error": err}}
            )
            raise HTTPException(status_code=502, detail=err)
    except _httpx.HTTPError as e:
        raise HTTPException(status_code=502, detail=str(e)[:120])

# Flow Builder
@router.get("/flows")
async def list_flows(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    flows = await db.flow_builders.find(
        {"company_id": user["company_id"]},
        {"_id": 0}
    ).to_list(1000)
    return flows

@router.post("/flows")
async def create_flow(
    data: FlowCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    flow = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "name": data.name,
        "description": data.description,
        "nodes": data.nodes,
        "edges": data.edges,
        "trigger_type": data.trigger_type,
        "is_active": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.flow_builders.insert_one(flow)
    return {k: v for k, v in flow.items() if k != "_id"}


@router.get("/flows/{flow_id}/export")
async def export_flow(
    flow_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    """Export a flow as a portable JSON. Strips tenant metadata so it can be
    re-imported into any company via /api/crm/flows/import.
    """
    flow = await db.flow_builders.find_one(
        {"id": flow_id, "company_id": user["company_id"]}, {"_id": 0}
    )
    if not flow:
        raise HTTPException(404, "Fluxo nao encontrado")
    out = {
        "name": flow.get("name"),
        "description": flow.get("description"),
        "trigger_type": flow.get("trigger_type"),
        "nodes": flow.get("nodes") or [],
        "edges": flow.get("edges") or [],
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "exported_from": "AgentCRM",
    }
    return out


@router.post("/flows/import")
async def import_flow(
    payload: dict,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    """Generic flow importer. Accepts the JSON exported from another tenant
    (this very app's `/api/crm/flows`) or any structurally compatible file.
    The frontend reads a .json from the user's computer and POSTs the parsed
    object here. We strip metadata that must NOT carry over (id, company_id,
    timestamps), keep nodes/edges, and force `is_active=False` so the admin
    can review before publishing.
    """
    if not isinstance(payload, dict):
        raise HTTPException(400, "JSON inválido — envie um objeto.")
    nodes = payload.get("nodes")
    edges = payload.get("edges")
    if not isinstance(nodes, list) or not isinstance(edges, list):
        raise HTTPException(400, "JSON inválido — esperado campos 'nodes' (lista) e 'edges' (lista). Exporte um fluxo deste sistema e tente novamente.")
    base_name = (payload.get("name") or "Fluxo importado").strip() or "Fluxo importado"
    # Collision-free name: append (N) if a flow with the same name already exists.
    name = base_name
    suffix = 1
    while await db.flow_builders.find_one({"company_id": user["company_id"], "name": name}, {"_id": 0, "id": 1}):
        suffix += 1
        name = f"{base_name} ({suffix})"
    flow = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "name": name,
        "description": (payload.get("description") or "Fluxo importado de arquivo JSON")[:500],
        "nodes": nodes,
        "edges": edges,
        "trigger_type": payload.get("trigger_type") or "manual",
        "is_active": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.flow_builders.insert_one(flow)
    return {k: v for k, v in flow.items() if k != "_id"}

@router.put("/flows/{flow_id}")
async def update_flow(
    flow_id: str,
    data: FlowUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    flow = await db.flow_builders.find_one({"id": flow_id, "company_id": user["company_id"]})
    if not flow:
        raise HTTPException(status_code=404, detail="Flow não encontrado")
    
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    
    if update_data:
        await db.flow_builders.update_one(
            {"id": flow_id},
            {"$set": update_data}
        )
    
    updated_flow = await db.flow_builders.find_one({"id": flow_id}, {"_id": 0})
    return updated_flow


@router.delete("/flows/{flow_id}")
async def delete_flow(
    flow_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    res = await db.flow_builders.delete_one({"id": flow_id, "company_id": user["company_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Flow nao encontrado")
    return {"message": "Flow removido"}


# === FLOW EXECUTION (auto-trigger when WhatsApp connection has default_flow_id) ===
async def _trigger_flow_for_ticket(db: AsyncIOMotorDatabase, company_id: str, flow_id: str, ticket: dict) -> None:
    """Kicks off the Flowbuilder runtime on a brand-new ticket. The runtime
    walks the flow graph, sending message/menu/http nodes as outgoing
    WhatsApp messages, and persists the current node id on the ticket so
    subsequent customer replies can advance the flow.
    See `flow_engine.py` for the full state machine.
    """
    flow = await db.flow_builders.find_one({"id": flow_id, "company_id": company_id}, {"_id": 0})
    if not flow or not (flow.get("nodes") or []):
        return
    from flow_engine import advance_flow
    await advance_flow(db, ticket, flow, is_initial=True)


# === Flow debug endpoints (admin/owner) — useful for troubleshooting why a
# ticket is "stuck" on a node in production. ===
@router.get("/tickets/{ticket_id}/flow-state")
async def get_flow_state(
    ticket_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    t = await db.tickets.find_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"_id": 0, "id": 1, "ticket_number": 1, "customer_name": 1, "customer_phone": 1,
         "connection_id": 1, "active_flow_id": 1, "active_flow_node_id": 1,
         "flow_started_at": 1, "flow_vars": 1, "status": 1},
    )
    if not t:
        raise HTTPException(404, "Ticket nao encontrado")
    flow_summary = None
    if t.get("active_flow_id"):
        f = await db.flow_builders.find_one({"id": t["active_flow_id"], "company_id": user["company_id"]}, {"_id": 0, "name": 1, "nodes": 1, "edges": 1})
        if f:
            cur_node = next((n for n in (f.get("nodes") or []) if n.get("id") == t.get("active_flow_node_id")), None)
            flow_summary = {
                "name": f.get("name"),
                "node_count": len(f.get("nodes") or []),
                "edge_count": len(f.get("edges") or []),
                "current_node": cur_node,
            }
    return {"ticket": t, "flow": flow_summary}


@router.post("/tickets/{ticket_id}/reset-flow")
async def reset_flow(
    ticket_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """Clear flow state on a ticket. Useful when a flow gets stuck or the
    flow definition has changed and the saved node id no longer exists."""
    res = await db.tickets.update_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"$set": {
            "active_flow_id": None,
            "active_flow_node_id": None,
            "flow_started_at": None,
            "flow_vars": {},
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Ticket nao encontrado")
    return {"ok": True, "modified": res.modified_count}


class _FlowTestPayload(BaseModel):
    incoming_text: Optional[str] = None
    is_initial: Optional[bool] = None


@router.post("/tickets/{ticket_id}/test-flow")
async def test_flow_advance(
    ticket_id: str,
    payload: _FlowTestPayload,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """DRY-RUN advance the flow against this ticket's current state. Returns
    the messages that WOULD be sent and the projected next node — without
    persisting state or pushing to WhatsApp.
    """
    t = await db.tickets.find_one({"id": ticket_id, "company_id": user["company_id"]}, {"_id": 0})
    if not t:
        raise HTTPException(404, "Ticket nao encontrado")
    flow_id = t.get("active_flow_id")
    is_initial = payload.is_initial if payload.is_initial is not None else (not flow_id)
    if not flow_id:
        # Try to get the connection's default flow
        if t.get("connection_id"):
            conn = await db.channel_connections.find_one({"id": t["connection_id"], "company_id": user["company_id"]}, {"_id": 0, "default_flow_id": 1})
            if conn and conn.get("default_flow_id"):
                flow_id = conn["default_flow_id"]
    if not flow_id:
        raise HTTPException(400, "Ticket nao tem fluxo ativo nem conexao com fluxo padrao")
    flow = await db.flow_builders.find_one({"id": flow_id, "company_id": user["company_id"]}, {"_id": 0})
    if not flow:
        raise HTTPException(404, "Fluxo nao encontrado")
    from flow_engine import advance_flow
    sent = await advance_flow(db, t, flow, incoming_text=payload.incoming_text, is_initial=is_initial, dry_run=True)
    return {"messages_sent": sent, "is_initial": is_initial, "flow_name": flow.get("name")}


# === TAGS ===
from pydantic import BaseModel as _BM


class TagCreate(_BM):
    name: str
    color: Optional[str] = "#64748B"
    description: Optional[str] = ""


class TagUpdate(_BM):
    name: Optional[str] = None
    color: Optional[str] = None
    description: Optional[str] = None


@router.get("/tags")
async def list_tags(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    tags = await db.tags.find(
        {"company_id": user["company_id"]},
        {"_id": 0}
    ).sort("name", 1).to_list(500)
    return tags


@router.post("/tags")
async def create_tag(
    data: TagCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    doc = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "name": data.name,
        "color": data.color or "#64748B",
        "description": data.description or "",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.tags.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@router.put("/tags/{tag_id}")
async def update_tag(
    tag_id: str,
    data: TagUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    update = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="Sem dados")
    res = await db.tags.update_one(
        {"id": tag_id, "company_id": user["company_id"]},
        {"$set": update}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tag nao encontrada")
    return await db.tags.find_one({"id": tag_id}, {"_id": 0})


@router.delete("/tags/{tag_id}")
async def delete_tag(
    tag_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    res = await db.tags.delete_one({"id": tag_id, "company_id": user["company_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tag nao encontrada")
    return {"message": "Tag removida"}


# === KANBAN COLUMNS ===
class KanbanColumnCreate(_BM):
    name: str
    color: Optional[str] = "#64748B"
    order: Optional[int] = 0


class KanbanColumnUpdate(_BM):
    name: Optional[str] = None
    color: Optional[str] = None
    order: Optional[int] = None


# A native first column ("Atendimentos") is always returned. It collects
# every ticket whose status doesn't match a custom column. Companies cannot
# delete or rename it from the UI.
NATIVE_FIRST_COLUMN = {
    "id": "native:atendimentos",
    "name": "Aguardando",
    "color": "#F59E0B",
    "order": 0,
    "is_native": True,
}


@router.get("/kanban-columns")
async def list_kanban_columns(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    custom = await db.kanban_columns.find(
        {"company_id": user["company_id"]},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    return [NATIVE_FIRST_COLUMN] + custom


@router.post("/kanban-columns")
async def create_kanban_column(
    data: KanbanColumnCreate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    # Auto-assign next order if not provided
    cnt = await db.kanban_columns.count_documents({"company_id": user["company_id"]})
    order = data.order if (data.order is not None and data.order > 0) else (cnt + 1)
    doc = {
        "id": str(uuid.uuid4()),
        "company_id": user["company_id"],
        "name": data.name,
        "color": data.color or "#64748B",
        "order": order,
        "is_native": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.kanban_columns.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@router.put("/kanban-columns/{column_id}")
async def update_kanban_column(
    column_id: str,
    data: KanbanColumnUpdate,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    if column_id.startswith("native:"):
        raise HTTPException(status_code=400, detail="Coluna nativa nao pode ser editada")
    update = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="Sem dados")
    res = await db.kanban_columns.update_one(
        {"id": column_id, "company_id": user["company_id"]},
        {"$set": update}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Coluna nao encontrada")
    return await db.kanban_columns.find_one({"id": column_id}, {"_id": 0})


@router.delete("/kanban-columns/{column_id}")
async def delete_kanban_column(
    column_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    if column_id.startswith("native:"):
        raise HTTPException(status_code=400, detail="Coluna nativa nao pode ser excluida")
    res = await db.kanban_columns.delete_one({"id": column_id, "company_id": user["company_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Coluna nao encontrada")
    return {"message": "Coluna removida"}


# Override kanban endpoint to use custom columns
@router.get("/kanban-v2")
async def get_kanban_v2(
    days: int = 90,
    search: Optional[str] = None,
    column_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    """Kanban grouped by company-defined columns (plus the native first one).

    Default window: last 90 days. Pass `days=0` to disable. Also accepts
    `search` (matches customer_name/phone/ticket_number) and `column_id`
    (limit to a single column — useful when the operator focuses on one
    column with thousands of tickets in it).
    """
    custom_cols = await db.kanban_columns.find(
        {"company_id": user["company_id"]}, {"_id": 0}
    ).sort("order", 1).to_list(100)
    columns = [NATIVE_FIRST_COLUMN] + custom_cols
    custom_ids = {c["id"] for c in custom_cols}

    query = {"company_id": user["company_id"]}
    # 2026-05-27 — `view_all_kanban` per-user flag bypasses queue/conn
    # filtering for the Kanban view (operator-tier "supervisao" sem mexer
    # no perfil de permissoes). Outras telas mantem o filtro normal.
    if user.get("view_all_kanban"):
        vis = {}
    else:
        vis = _ticket_visibility_filter(user)
    if vis:
        query.update(vis)

    if days and days > 0:
        cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
        # Tickets with NO `updated_at` (legacy) are kept visible to avoid
        # silently hiding old data.
        query["$or"] = (query.get("$or") or []) + [
            {"updated_at": {"$gte": cutoff}},
            {"created_at": {"$gte": cutoff}},
        ] if not query.get("$or") else query.get("$or")
        # If we already had an $or, AND-combine with the time window:
        if "$or" in query and not query.get("$and"):
            # rebuild as $and to combine cleanly
            time_clause = {"$or": [{"updated_at": {"$gte": cutoff}}, {"created_at": {"$gte": cutoff}}]}
            vis_or = query.pop("$or")
            query["$and"] = [{"$or": vis_or}, time_clause]
        else:
            query.setdefault("$and", []).append({"$or": [{"updated_at": {"$gte": cutoff}}, {"created_at": {"$gte": cutoff}}]})

    if search:
        s = search.strip()
        digits = re.sub(r"\D", "", s)
        sr = []
        sr.append({"customer_name": {"$regex": s, "$options": "i"}})
        if digits:
            sr.append({"customer_phone": {"$regex": digits}})
            sr.append({"ticket_number": int(digits) if digits.isdigit() else digits})
        query.setdefault("$and", []).append({"$or": sr})

    if column_id and column_id != NATIVE_FIRST_COLUMN["id"]:
        query["kanban_column_id"] = column_id
    tickets = await db.tickets.find(query, {"_id": 0, "messages": 0}).sort("updated_at", -1).to_list(2000)

    grouped = {c["id"]: [] for c in columns}
    for t in tickets:
        col = t.get("kanban_column_id")
        if col and col in custom_ids:
            grouped[col].append(t)
        else:
            grouped[NATIVE_FIRST_COLUMN["id"]].append(t)

    totals_by_column = {
        col_id: sum(float(t.get("value") or 0) for t in items)
        for col_id, items in grouped.items()
    }

    return {"columns": columns, "tickets_by_column": grouped, "totals_by_column": totals_by_column, "window_days": days}


class KanbanReorderRequest(BaseModel):
    column_ids: List[str]  # ordered list of column ids; first one shown left-most


@router.post("/kanban-columns/reorder")
async def reorder_kanban_columns(
    body: KanbanReorderRequest,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    """Persist a new column ordering. Native columns (`native:*`) are
    silently dropped from the input — they are always anchored on the
    left. Operators access this via a "disfarçado" button in the Kanban
    page header (long-press / admin-only)."""
    if not body.column_ids:
        raise HTTPException(400, "column_ids vazio")
    custom_ids = [c for c in body.column_ids if not c.startswith("native:")]
    # Bulk-update order (1-indexed; native column always sits at 0)
    for idx, col_id in enumerate(custom_ids, start=1):
        await db.kanban_columns.update_one(
            {"id": col_id, "company_id": user["company_id"]},
            {"$set": {"order": idx, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    return {"reordered": len(custom_ids)}


@router.put("/tickets/{ticket_id}/kanban-column")
async def move_ticket_to_column(
    ticket_id: str,
    body: dict,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database)
):
    column_id = body.get("column_id")
    if not column_id:
        raise HTTPException(status_code=400, detail="column_id obrigatorio")
    set_value: Any = column_id
    if column_id.startswith("native:"):
        # Native column: clear custom assignment so the ticket falls back here
        set_value = None
    res = await db.tickets.update_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"$set": {"kanban_column_id": set_value, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket nao encontrado")
    return {"message": "Ticket movido"}



# === BULK IMPORT (XLSX) =====================================================
def _norm(s: str) -> str:
    """Lowercase + strip + collapse spaces — used to compare tag/column names."""
    return re.sub(r"\s+", " ", (s or "").strip().lower())


@router.post("/clients/normalize-birth-dates")
async def normalize_birth_dates(
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """One-shot migration: convert non-ISO `birth_date` values currently in
    `db.clients` to ISO `YYYY-MM-DD`. Required after the 04/05/2026 fix —
    contacts imported with a BR-formatted date (DD/MM/YYYY) were stored as
    a free string and never rendered in the UI. Idempotent: rows already in
    ISO are left untouched. Restricted to admin/owner of the calling
    company; only normalizes clients OF THAT company."""
    role = (user.get("role") or "").lower()
    if role not in ("super_admin", "superadmin", "company_admin", "owner"):
        raise HTTPException(status_code=403, detail="Apenas administradores podem executar esta migracao")

    company_id = user["company_id"]
    cursor = db.clients.find(
        {"company_id": company_id, "birth_date": {"$exists": True, "$ne": None, "$ne": ""}},
        {"_id": 0, "id": 1, "birth_date": 1},
    )
    converted = 0
    skipped_already_iso = 0
    failed: list[dict] = []
    async for c in cursor:
        cur = c.get("birth_date")
        if not isinstance(cur, str):
            continue
        s = cur.strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}", s):
            skipped_already_iso += 1
            continue
        new_iso = None
        # BR format DD/MM/YYYY (or with -)
        m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", s)
        if m:
            d, mo, y = m.group(1), m.group(2), m.group(3)
            if len(y) == 2:
                y = ("19" + y) if int(y) > 30 else ("20" + y)
            try:
                from datetime import date as _date
                new_iso = _date(int(y), int(mo), int(d)).isoformat()
            except Exception:
                pass
        if not new_iso:
            try:
                import pandas as _pd
                ts = _pd.to_datetime(s, dayfirst=True, errors="raise")
                new_iso = ts.date().isoformat()
            except Exception:
                failed.append({"id": c["id"], "value": s})
                continue
        await db.clients.update_one(
            {"id": c["id"], "company_id": company_id},
            {"$set": {"birth_date": new_iso}},
        )
        converted += 1
    return {
        "converted": converted,
        "skipped_already_iso": skipped_already_iso,
        "failed_count": len(failed),
        "failed_sample": failed[:30],
    }


@router.get("/clients/import-xlsx-template")
async def download_clients_import_template(user: dict = Depends(get_current_user)):
    """Return a ready-to-fill `.xlsx` template with the columns the importer
    accepts. The first sheet has 2 example rows (PJ + PF) so the operator
    sees how to format `data de nascimento`, `tags e Kambam`, CPF/CNPJ etc.
    A second sheet documents each column."""
    import pandas as pd
    columns = [
        "name", "Telefone", "email", "data de nascimento",
        "tipo de pessoa", "cpf", "cnpj", "razao social",
        "cep", "endereco", "cidade", "estado",
        "tags e Kambam", "observacoes",
    ]
    examples = [
        {
            "name": "Joao Silva",
            "Telefone": "5511988887777",
            "email": "joao@example.com",
            "data de nascimento": "1990-05-12",
            "tipo de pessoa": "fisica",
            "cpf": "123.456.789-00",
            "cnpj": "",
            "razao social": "",
            "cep": "01310-100",
            "endereco": "Av. Paulista, 1000",
            "cidade": "Sao Paulo",
            "estado": "SP",
            "tags e Kambam": "Novo Cliente, PROSPECTAR",
            "observacoes": "Indicado por Maria",
        },
        {
            "name": "ACME Industria LTDA",
            "Telefone": "5562999998888",
            "email": "contato@acme.com.br",
            "data de nascimento": "",
            "tipo de pessoa": "juridica",
            "cpf": "",
            "cnpj": "07.393.407/0001-75",
            "razao social": "ACME Industria LTDA",
            "cep": "75252-320",
            "endereco": "Rua das Flores, 200",
            "cidade": "Senador Canedo",
            "estado": "GO",
            "tags e Kambam": "WON - PROP. FECHADA",
            "observacoes": "",
        },
    ]
    docs = [
        {"coluna": "name",                "obrigatorio": "SIM",  "descricao": "Nome do cliente / razao social abreviada"},
        {"coluna": "Telefone",            "obrigatorio": "SIM",  "descricao": "Com DDI 55, somente digitos. Ex.: 5511988887777"},
        {"coluna": "email",               "obrigatorio": "nao",  "descricao": "E-mail do contato"},
        {"coluna": "data de nascimento",  "obrigatorio": "nao",  "descricao": "ISO YYYY-MM-DD ou data do Excel. Ex.: 1990-05-12"},
        {"coluna": "tipo de pessoa",      "obrigatorio": "nao",  "descricao": "fisica | juridica (default: fisica, ou juridica se houver CNPJ)"},
        {"coluna": "cpf",                 "obrigatorio": "nao",  "descricao": "CPF para PF"},
        {"coluna": "cnpj",                "obrigatorio": "nao",  "descricao": "CNPJ para PJ"},
        {"coluna": "razao social",        "obrigatorio": "nao",  "descricao": "Razao social / nome da empresa (PJ)"},
        {"coluna": "cep",                 "obrigatorio": "nao",  "descricao": "CEP do endereco"},
        {"coluna": "endereco",            "obrigatorio": "nao",  "descricao": "Logradouro completo"},
        {"coluna": "cidade",              "obrigatorio": "nao",  "descricao": "Cidade"},
        {"coluna": "estado",              "obrigatorio": "nao",  "descricao": "UF (2 letras). Ex.: SP, GO"},
        {"coluna": "tags e Kambam",       "obrigatorio": "nao",  "descricao": "Lista separada por virgula. Cada item bate com uma Tag ou Coluna do Kanban da empresa (case-insensitive)."},
        {"coluna": "observacoes",         "obrigatorio": "nao",  "descricao": "Texto livre que vai para o campo Observacoes do cliente"},
    ]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(examples, columns=columns).to_excel(
            writer, index=False, sheet_name="clientes"
        )
        pd.DataFrame(docs).to_excel(writer, index=False, sheet_name="instrucoes")
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo-importacao-clientes.xlsx"'},
    )


@router.post("/clients/import-xlsx")
async def import_clients_xlsx(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """One-shot import of a `.xlsx` contact backup into the company.

    Expected columns (case-insensitive): `name`, `Telefone`, `email`,
    `tags e Kambam`. Each entry in the last column is matched against the
    company's existing `tags` (added to client.tags) or `kanban_columns`
    (a single ticket is created/updated to anchor the client in the kanban
    board on the latest matching column). Duplicates by phone (digits-only)
    are merged: name/email/tags get refreshed from the file.

    Restricted to owners/admins of the calling company.
    """
    role = (user.get("role") or "").lower()
    if role not in ("super_admin", "superadmin", "company_admin", "owner"):
        raise HTTPException(status_code=403, detail="Apenas administradores podem importar contatos")

    company_id = user["company_id"]
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Arquivo vazio")

    try:
        import pandas as pd  # local import keeps cold-start light
        df = pd.read_excel(io.BytesIO(raw))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Falha ao ler XLSX: {e}")

    # Normalize headers — accept upper/lowercase + a few aliases (PT/EN)
    cols = {c.lower().strip(): c for c in df.columns}
    def pick(*aliases):
        for a in aliases:
            if a in cols:
                return cols[a]
        return None
    name_col         = pick("name", "nome")
    phone_col        = pick("telefone", "phone", "celular")
    email_col        = pick("email", "e-mail")
    tags_col         = pick("tags e kambam", "tags", "tags e kanban", "tag")
    birth_date_col   = pick("data de nascimento", "data_nascimento", "birth_date", "nascimento")
    person_type_col  = pick("tipo de pessoa", "person_type", "tipo")
    cpf_col          = pick("cpf")
    cnpj_col         = pick("cnpj")
    company_name_col = pick("razao social", "razão social", "company_name", "nome da empresa", "empresa")
    cep_col          = pick("cep")
    address_col      = pick("endereco", "endereço", "address")
    city_col         = pick("cidade", "city")
    state_col        = pick("estado", "state", "uf")
    notes_col        = pick("observacoes", "observações", "notes", "obs")

    if not (name_col and phone_col):
        raise HTTPException(
            status_code=400,
            detail=f"Colunas obrigatorias ausentes (esperado: name, Telefone). Encontradas: {list(df.columns)}",
        )

    # Pre-load the company tags + kanban columns so we can match by name
    tags_docs = await db.tags.find({"company_id": company_id}, {"_id": 0}).to_list(2000)
    cols_docs = await db.kanban_columns.find({"company_id": company_id}, {"_id": 0}).to_list(500)
    tags_by_name = {_norm(t["name"]): t for t in tags_docs}
    cols_by_name = {_norm(c["name"]): c for c in cols_docs}

    # Cache existing clients (digits-only phone -> client doc)
    existing = {}
    async for c in db.clients.find({"company_id": company_id}, {"_id": 0}):
        d = normalize_phone(c.get("phone"))
        if d:
            existing[d] = c

    created = 0
    updated = 0
    skipped_no_phone = 0
    tickets_created = 0
    tickets_updated = 0
    unknown_labels: dict[str, int] = {}

    now_iso = datetime.now(timezone.utc).isoformat()

    for _, row in df.iterrows():
        name = str(row[name_col]).strip() if row[name_col] is not None and str(row[name_col]) != "nan" else ""
        phone_raw = "" if row[phone_col] is None else str(row[phone_col]).strip()
        # pandas may store ints — kill the trailing `.0`
        if phone_raw.endswith(".0"):
            phone_raw = phone_raw[:-2]
        digits = normalize_phone(phone_raw)
        if not digits:
            skipped_no_phone += 1
            continue

        email_val = None
        if email_col is not None:
            ev = row[email_col]
            if ev is not None and str(ev) != "nan" and str(ev).strip():
                email_val = str(ev).strip()

        # Optional extra fields. We trust the spreadsheet — admins can clean
        # up later from the contact panel. `birth_date` is normalized to
        # ISO YYYY-MM-DD when pandas parsed it as a Timestamp.
        def cell(col):
            if col is None:
                return None
            v = row[col]
            if v is None:
                return None
            s = str(v).strip()
            if not s or s.lower() == "nan":
                return None
            return s

        birth_val = None
        if birth_date_col is not None:
            v = row[birth_date_col]
            if v is not None and str(v).strip() and str(v).lower() != "nan":
                # Normalize to ISO YYYY-MM-DD so the frontend can render it
                # via `new Date(...)`. We accept:
                #   * pandas Timestamp (when Excel cell was a real date)
                #   * BR string DD/MM/YYYY or DD-MM-YYYY (most common)
                #   * ISO string YYYY-MM-DD (already normalized)
                # Anything else is stored verbatim as a last-resort.
                try:
                    import pandas as _pd
                    if isinstance(v, _pd.Timestamp):
                        birth_val = v.date().isoformat()
                    else:
                        s = str(v).strip()
                        # try BR format DD/MM/YYYY (or DD-MM-YYYY)
                        m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", s)
                        if m:
                            d, mo, y = m.group(1), m.group(2), m.group(3)
                            if len(y) == 2:
                                y = ("19" + y) if int(y) > 30 else ("20" + y)
                            try:
                                from datetime import date as _date
                                birth_val = _date(int(y), int(mo), int(d)).isoformat()
                            except Exception:
                                birth_val = s
                        elif re.match(r"^\d{4}-\d{2}-\d{2}", s):
                            birth_val = s[:10]
                        else:
                            # try pandas as last resort
                            try:
                                ts = _pd.to_datetime(s, dayfirst=True, errors="raise")
                                birth_val = ts.date().isoformat()
                            except Exception:
                                birth_val = s
                except Exception:
                    birth_val = str(v).strip()

        person_type_val = (cell(person_type_col) or "").lower()
        if person_type_val not in ("fisica", "juridica"):
            person_type_val = "juridica" if cell(cnpj_col) else "fisica"

        extras = {
            "person_type": person_type_val,
            "cpf": cell(cpf_col),
            "cnpj": cell(cnpj_col),
            "company_name": cell(company_name_col),
            "cep": cell(cep_col),
            "address": cell(address_col),
            "city": cell(city_col),
            "state": cell(state_col),
            "notes": cell(notes_col),
            "birth_date": birth_val,
        }
        # Drop None fields so we don't overwrite existing data with blanks
        extras_nonnull = {k: v for k, v in extras.items() if v}

        labels: list[str] = []
        if tags_col is not None:
            tv = row[tags_col]
            if tv is not None and str(tv) != "nan":
                labels = [s.strip() for s in str(tv).split(",") if s.strip()]

        # Split labels into tag-names vs kanban-column matches
        matched_tags: list[str] = []      # tag names (will land in client.tags)
        matched_columns: list[dict] = []  # kanban_columns docs
        for lbl in labels:
            key = _norm(lbl)
            if key in tags_by_name:
                matched_tags.append(tags_by_name[key]["name"])
            elif key in cols_by_name:
                matched_columns.append(cols_by_name[key])
            else:
                unknown_labels[lbl] = unknown_labels.get(lbl, 0) + 1
                # Default behaviour: still keep the label as a free-form tag so
                # the data isn't lost — admins can clean it up later.
                matched_tags.append(lbl)

        # Pick the LAST kanban column listed as the anchor (most-recent stage)
        anchor_col = matched_columns[-1] if matched_columns else None

        existing_client = existing.get(digits)
        if existing_client:
            # Merge: union tags, refresh name/email if non-empty
            cur_tags = list(existing_client.get("tags") or [])
            merged_tags = cur_tags[:]
            for t in matched_tags:
                if t not in merged_tags:
                    merged_tags.append(t)
            update_set: dict = {"tags": merged_tags, "updated_at": now_iso}
            if name:
                update_set["name"] = name
            if email_val:
                update_set["email"] = email_val
            update_set.update(extras_nonnull)
            await db.clients.update_one(
                {"id": existing_client["id"], "company_id": company_id},
                {"$set": update_set},
            )
            client_id = existing_client["id"]
            client_name = name or existing_client.get("name") or phone_raw
            client_phone = existing_client.get("phone") or phone_raw
            updated += 1
        else:
            client_id = str(uuid.uuid4())
            doc = {
                "id": client_id,
                "company_id": company_id,
                "name": name or phone_raw,
                "phone": phone_raw,
                "email": email_val,
                "person_type": person_type_val or "fisica",
                "tags": matched_tags,
                "total_appointments": 0,
                "created_at": now_iso,
                "created_via": "xlsx_import",
                **extras_nonnull,
            }
            await db.clients.insert_one(doc)
            existing[digits] = {**doc}
            client_name = name or phone_raw
            client_phone = phone_raw
            created += 1

        # Anchor in kanban (only when at least one matched column)
        if anchor_col:
            t_existing = await db.tickets.find_one(
                {"company_id": company_id, "client_id": client_id},
                {"_id": 0, "id": 1},
            )
            if t_existing:
                await db.tickets.update_one(
                    {"id": t_existing["id"]},
                    {"$set": {
                        "kanban_column_id": anchor_col["id"],
                        "customer_name": client_name,
                        "customer_phone": client_phone,
                        "updated_at": now_iso,
                    }},
                )
                tickets_updated += 1
            else:
                t_id = str(uuid.uuid4())
                tnum = await next_ticket_number(db, company_id)
                await db.tickets.insert_one({
                    "id": t_id,
                    "ticket_number": tnum,
                    "company_id": company_id,
                    "client_id": client_id,
                    "customer_name": client_name,
                    "customer_phone": client_phone,
                    "customer_email": email_val,
                    "status": "aberto",
                    "priority": "media",
                    "channel": "import",
                    "description": "Importado via XLSX",
                    "assigned_to": None,
                    "messages": [],
                    "tags": [],
                    "value": 0.0,
                    "kanban_column_id": anchor_col["id"],
                    "created_at": now_iso,
                    "updated_at": now_iso,
                })
                tickets_created += 1

    # Top-N unknown labels for the report
    top_unknown = sorted(unknown_labels.items(), key=lambda x: -x[1])[:30]

    return {
        "rows_total": int(len(df)),
        "created": created,
        "updated": updated,
        "skipped_no_phone": skipped_no_phone,
        "tickets_created": tickets_created,
        "tickets_updated": tickets_updated,
        "unknown_labels_count": len(unknown_labels),
        "unknown_labels_top": [{"label": k, "count": v} for k, v in top_unknown],
    }



# ──────────────────────────────────────────────────────────────────────
# 2026-02-18 — Editar / Apagar mensagem ja enviada via WhatsApp
# ──────────────────────────────────────────────────────────────────────
async def _wa_service_url() -> str:
    return os.environ.get("WA_SERVICE_URL", "http://localhost:3002")


@router.post("/tickets/{ticket_id}/messages/{message_id}/edit")
async def edit_outbound_message(
    ticket_id: str,
    message_id: str,
    payload: dict,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    new_text = (payload or {}).get("text", "").strip()
    if not new_text:
        raise HTTPException(400, "text obrigatorio")
    ticket = await db.tickets.find_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"_id": 0, "messages": 1, "customer_phone": 1, "connection_id": 1},
    )
    if not ticket:
        raise HTTPException(404, "Ticket nao encontrado")
    msg = next((m for m in (ticket.get("messages") or []) if m.get("id") == message_id), None)
    if not msg:
        raise HTTPException(404, "Mensagem nao encontrada")
    if msg.get("sender_type") != "agent":
        raise HTTPException(400, "So mensagens enviadas pelo operador podem ser editadas")
    if not msg.get("wa_message_id"):
        raise HTTPException(400, "Mensagem ainda nao tem ID do WhatsApp — aguarde a confirmacao do envio")
    if msg.get("deleted_for_customer"):
        raise HTTPException(400, "Mensagem ja apagada, nao pode ser editada")
    import httpx
    wa_url = await _wa_service_url()
    wa_error = None
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            r = await client.post(
                f"{wa_url}/instances/{ticket['connection_id']}/edit-message",
                json={
                    "phone": ticket["customer_phone"],
                    "message_id": msg["wa_message_id"],
                    "new_text": new_text,
                },
            )
            if r.status_code >= 400:
                _ct = r.headers.get("content-type", "")
                detail = (r.json().get("error") if _ct.startswith("application/json") else r.text)[:200]
                # 2026-02-18 — Para EDIT, divergencia eh mais grave (texto local
                # ≠ WhatsApp). Aqui ainda devolvemos 400 para o operador saber.
                raise HTTPException(r.status_code, f"WhatsApp recusou edicao: {detail}")
    except httpx.RequestError as e:
        # Mesmo padrao do delete: o Baileys pode demorar mas concluir.
        wa_error = f"timeout/network: {str(e)[:120]}"
    now_iso = datetime.now(timezone.utc).isoformat()
    edit_entry = {
        "previous_text": msg.get("content") or "",
        "edited_at": now_iso,
        "edited_by_id": user.get("id"),
        "edited_by_name": user.get("name") or user.get("email") or "",
    }
    await db.tickets.update_one(
        {"id": ticket_id, "messages.id": message_id},
        {
            "$set": {
                "messages.$.content": new_text,
                "messages.$.edited_at": now_iso,
                **({"messages.$.wa_edit_error": wa_error} if wa_error else {}),
            },
            "$push": {"messages.$.edit_history": edit_entry},
        },
    )
    return {"ok": True, "message_id": message_id, "new_text": new_text, "warning": wa_error}


@router.post("/tickets/{ticket_id}/messages/{message_id}/delete-for-customer")
async def delete_outbound_message(
    ticket_id: str,
    message_id: str,
    user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    ticket = await db.tickets.find_one(
        {"id": ticket_id, "company_id": user["company_id"]},
        {"_id": 0, "messages": 1, "customer_phone": 1, "connection_id": 1},
    )
    if not ticket:
        raise HTTPException(404, "Ticket nao encontrado")
    msg = next((m for m in (ticket.get("messages") or []) if m.get("id") == message_id), None)
    if not msg:
        raise HTTPException(404, "Mensagem nao encontrada")
    if msg.get("sender_type") != "agent":
        raise HTTPException(400, "So mensagens enviadas pelo operador podem ser apagadas")
    if not msg.get("wa_message_id"):
        raise HTTPException(400, "Mensagem ainda nao tem ID do WhatsApp")
    if msg.get("deleted_for_customer"):
        return {"ok": True, "already": True}
    import httpx
    wa_url = await _wa_service_url()
    wa_ok = False
    wa_error = None
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            r = await client.post(
                f"{wa_url}/instances/{ticket['connection_id']}/delete-message",
                json={
                    "phone": ticket["customer_phone"],
                    "message_id": msg["wa_message_id"],
                },
            )
            wa_ok = r.status_code < 400
            if not wa_ok:
                _ct = r.headers.get("content-type", "")
                wa_error = (r.json().get("error") if _ct.startswith("application/json") else r.text)[:200]
    except httpx.RequestError as e:
        # 2026-02-18 — O Baileys frequentemente revoga a mensagem no WhatsApp
        # mas demora a responder por sobrecarga, gerando timeout aqui. Nao
        # devolvemos erro — marcamos local com flag wa_delete_error pra
        # auditoria. Operacao concluida do ponto de vista do cliente final.
        wa_error = f"timeout/network: {str(e)[:120]}"
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.tickets.update_one(
        {"id": ticket_id, "messages.id": message_id},
        {"$set": {
            "messages.$.deleted_for_customer": True,
            "messages.$.deleted_at": now_iso,
            "messages.$.deleted_by_id": user.get("id"),
            "messages.$.deleted_by_name": user.get("name") or user.get("email") or "",
            **({"messages.$.wa_delete_error": wa_error} if wa_error else {}),
        }},
    )
    return {"ok": True, "message_id": message_id, "wa_ok": wa_ok, "warning": wa_error}

