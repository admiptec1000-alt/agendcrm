"""Iteration 60 tests — Campaign pause/resume/cancel/progress, Excel
import/template, and scheduler `_process_scheduled_campaigns` for
status='programada' campaigns.
"""
import os
import io
import uuid
import asyncio
import pytest
import requests
from datetime import datetime, timezone, timedelta

def _load_frontend_env():
    try:
        with open("/app/frontend/.env") as f:
            for ln in f:
                if ln.startswith("REACT_APP_BACKEND_URL="):
                    return ln.split("=", 1)[1].strip()
    except Exception:
        pass
    return os.environ.get("REACT_APP_BACKEND_URL", "")


BASE_URL = (os.environ.get("REACT_APP_BACKEND_URL") or _load_frontend_env()).rstrip("/")
API = f"{BASE_URL}/api"
CRM_EMAIL = "crm@test.com"
CRM_PASS = "crm123"


@pytest.fixture(scope="module")
def crm_token():
    r = requests.post(f"{API}/auth/login", json={"email": CRM_EMAIL, "password": CRM_PASS}, timeout=15)
    assert r.status_code == 200, f"CRM login failed: {r.status_code} {r.text}"
    data = r.json()
    return data.get("access_token") or data.get("token")


@pytest.fixture(scope="module")
def headers(crm_token):
    return {"Authorization": f"Bearer {crm_token}"}


@pytest.fixture(scope="module")
def company_id(crm_token):
    # Decode from the token via a sample call
    h = {"Authorization": f"Bearer {crm_token}"}
    r = requests.get(f"{API}/crm/campaigns", headers=h, timeout=10)
    # Just need a valid auth call; we don't actually use company_id directly
    return None


# ─── Excel template + import ─────────────────────────────────────────
class TestContactListExcel:
    def test_template_download(self, headers):
        r = requests.get(f"{API}/crm/contact-lists/template.xlsx", headers=headers, timeout=10)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml.sheet" in ct, f"bad content-type {ct}"
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower()
        assert ".xlsx" in cd.lower()
        # Verify it's a valid xlsx and has 'Contatos' sheet with right headers
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        assert "Contatos" in wb.sheetnames
        ws = wb["Contatos"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "Nome"
        assert rows[0][1] == "Telefone"
        assert rows[0][2] == "Email"

    def test_import_excel_rejects_csv(self, headers):
        files = {"file": ("contacts.csv", b"Nome,Telefone\nJoao,11999\n", "text/csv")}
        r = requests.post(
            f"{API}/crm/contact-lists/import-excel?name=TEST_csv_reject",
            files=files, headers=headers, timeout=15,
        )
        assert r.status_code == 400, f"expected 400, got {r.status_code} {r.text}"

    def test_import_excel_creates_list_with_dedup(self, headers):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Contatos"
        ws.append(["Nome", "Telefone", "Email"])
        ws.append(["Alice", "5511900000001", "a@a.com"])
        ws.append(["Bob",   "5511900000002", "b@b.com"])
        ws.append(["Alice2", "5511900000001", "a2@a.com"])  # duplicate -> skipped
        ws.append(["NoPhone", "", ""])  # missing phone -> skipped
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("TEST_import.xlsx", buf.read(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(
            f"{API}/crm/contact-lists/import-excel?name=TEST_iter60_list",
            files=files, headers=headers, timeout=20,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert "id" in body
        assert body.get("name") == "TEST_iter60_list"
        assert body.get("imported_count") == 2, f"imported_count={body.get('imported_count')}"
        assert body.get("skipped_count", 0) >= 1
        # Cleanup
        try:
            requests.delete(f"{API}/crm/contact-lists/{body['id']}", headers=headers, timeout=10)
        except Exception:
            pass


# ─── Campaign pause/resume/cancel/progress ───────────────────────────
@pytest.fixture(scope="module")
def created_campaign(headers):
    """Create a campaign with audiencia=[] for the empty-audience test."""
    payload = {
        "name": f"TEST_iter60_empty_{uuid.uuid4().hex[:6]}",
        "messages": ["Ola {nome}"],
        "audiencia": [],
        "status": "rascunho",
    }
    r = requests.post(f"{API}/crm/campaigns", json=payload, headers=headers, timeout=15)
    assert r.status_code in (200, 201), r.text
    camp = r.json()
    yield camp
    try:
        requests.delete(f"{API}/crm/campaigns/{camp['id']}", headers=headers, timeout=10)
    except Exception:
        pass


class TestCampaignControls:
    def test_run_with_empty_audience_returns_400(self, headers, created_campaign):
        r = requests.post(
            f"{API}/crm/campaigns/{created_campaign['id']}/run",
            headers=headers, timeout=15,
        )
        assert r.status_code == 400, f"expected 400, got {r.status_code} {r.text}"
        assert "udien" in r.text.lower() or "audien" in r.text.lower()

    def test_pause_nonrunning_returns_400(self, headers, created_campaign):
        r = requests.post(
            f"{API}/crm/campaigns/{created_campaign['id']}/pause",
            headers=headers, timeout=10,
        )
        assert r.status_code == 400

    def test_resume_nonpaused_returns_400(self, headers, created_campaign):
        r = requests.post(
            f"{API}/crm/campaigns/{created_campaign['id']}/resume",
            headers=headers, timeout=10,
        )
        assert r.status_code == 400

    def test_progress_shape(self, headers, created_campaign):
        r = requests.get(
            f"{API}/crm/campaigns/{created_campaign['id']}/progress",
            headers=headers, timeout=10,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert "campaign" in body
        assert "totals" in body
        for k in ("pending", "sending", "sent", "failed", "total"):
            assert k in body["totals"], f"missing totals.{k}"
        assert isinstance(body.get("sent"), list)
        assert isinstance(body.get("failed"), list)
        assert isinstance(body.get("pending"), list)
        assert isinstance(body.get("sending"), list)

    def test_cancel_endpoint(self, headers, created_campaign):
        # Force campaign into em_execucao via direct DB update? Use the
        # /run flow won't work since audience empty. Just test the 400
        # shape for an undraftable status.
        r = requests.post(
            f"{API}/crm/campaigns/{created_campaign['id']}/cancel",
            headers=headers, timeout=10,
        )
        # rascunho not in allowed list -> 400
        assert r.status_code == 400


# ─── Campaign with real audience (1 contact) ─────────────────────────
class TestCampaignRunWithAudience:
    @pytest.fixture(scope="class")
    def campaign_with_audience(self, headers):
        # Need a fake WA connection for this tenant; seed via direct Mongo
        import asyncio as _aio
        async def _seed():
            import sys; sys.path.insert(0, "/app/backend")
            from database import connect_to_mongo, get_database
            await connect_to_mongo()
            _db = await get_database()
            # Find tenant company_id from the crm@test.com user
            u = await _db.company_users.find_one({"email": CRM_EMAIL}, {"_id": 0, "company_id": 1})
            assert u, "crm@test.com user not found"
            cid = u["company_id"]
            conn_id = f"TEST_iter60_conn_{uuid.uuid4().hex[:6]}"
            await _db.channel_connections.insert_one({
                "id": conn_id, "company_id": cid, "type": "whatsapp",
                "status": "connected", "name": "TEST_iter60_fake",
            })
            return cid, conn_id, _db
        cid, conn_id, _db = _aio.get_event_loop().run_until_complete(_seed())

        # Create a contact list with one phone via Excel import
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Contatos"
        ws.append(["Nome", "Telefone", "Email"])
        ws.append(["TEST_recipient", "5511900000099", ""])
        ws.append(["TEST_recipient2", "5511900000098", ""])
        ws.append(["TEST_recipient3", "5511900000097", ""])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        files = {"file": ("c.xlsx", buf.read(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(
            f"{API}/crm/contact-lists/import-excel?name=TEST_iter60_camp_list",
            files=files, headers=headers, timeout=15,
        )
        assert r.status_code == 200, r.text
        list_id = r.json()["id"]

        payload = {
            "name": f"TEST_iter60_run_{uuid.uuid4().hex[:6]}",
            "messages": ["Ola"],
            "audience_mode": "list",
            "contact_list_id": list_id,
            "connection_id": conn_id,
            "status": "rascunho",
            "anti_block": {"enabled": True, "interval_min_seconds": 60,
                           "interval_max_seconds": 90, "burst_size": 50,
                           "burst_pause_seconds": 30, "daily_limit": 10},
        }
        r = requests.post(f"{API}/crm/campaigns", json=payload, headers=headers, timeout=15)
        assert r.status_code in (200, 201), r.text
        camp = r.json()
        yield camp
        # Cleanup
        async def _cleanup():
            await _db.channel_connections.delete_one({"id": conn_id})
            await _db.campaign_deliveries.delete_many({"campaign_id": camp["id"]})
        try:
            _aio.get_event_loop().run_until_complete(_cleanup())
        except Exception:
            pass
        try:
            requests.post(f"{API}/crm/campaigns/{camp['id']}/cancel", headers=headers, timeout=5)
        except Exception:
            pass
        try:
            requests.delete(f"{API}/crm/campaigns/{camp['id']}", headers=headers, timeout=10)
        except Exception:
            pass
        try:
            requests.delete(f"{API}/crm/contact-lists/{list_id}", headers=headers, timeout=10)
        except Exception:
            pass

    def test_run_seeds_deliveries(self, headers, campaign_with_audience):
        cid = campaign_with_audience["id"]
        r = requests.post(f"{API}/crm/campaigns/{cid}/run", headers=headers, timeout=20)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body.get("mode") == "classic"
        assert body.get("queued") is True
        assert body.get("total") == 3
        # Give it a moment to seed deliveries
        import time
        time.sleep(1)
        pr = requests.get(f"{API}/crm/campaigns/{cid}/progress", headers=headers, timeout=10)
        assert pr.status_code == 200
        totals = pr.json()["totals"]
        assert totals["total"] >= 1, f"deliveries not seeded: {totals}"

    def test_pause_during_execution(self, headers, campaign_with_audience):
        cid = campaign_with_audience["id"]
        # Give runner enough time to enter sleep state after first delivery
        import time
        time.sleep(3)
        r = requests.post(f"{API}/crm/campaigns/{cid}/pause", headers=headers, timeout=10)
        # Could be concluida already if only 1 delivery and very fast
        if r.status_code == 400:
            pr = requests.get(f"{API}/crm/campaigns/{cid}/progress", headers=headers, timeout=10)
            st = pr.json()["campaign"]["status"]
            assert st in ("concluida", "cancelada"), f"unexpected status {st}"
            pytest.skip(f"campaign finished before pause: status={st}")
        assert r.status_code == 200, r.text
        pr = requests.get(f"{API}/crm/campaigns/{cid}/progress", headers=headers, timeout=10)
        assert pr.json()["campaign"]["status"] == "pausada"

    def test_resume_paused(self, headers, campaign_with_audience):
        cid = campaign_with_audience["id"]
        # First ensure campaign is paused (if previous test skipped, force pause manually)
        pr = requests.get(f"{API}/crm/campaigns/{cid}/progress", headers=headers, timeout=10)
        if pr.json()["campaign"]["status"] != "pausada":
            pytest.skip(f"campaign not in pausada state: {pr.json()['campaign']['status']}")
        r = requests.post(f"{API}/crm/campaigns/{cid}/resume", headers=headers, timeout=10)
        assert r.status_code == 200, r.text
        pr = requests.get(f"{API}/crm/campaigns/{cid}/progress", headers=headers, timeout=10)
        assert pr.json()["campaign"]["status"] in ("em_execucao", "concluida")


# ─── Scheduler: _process_scheduled_campaigns ─────────────────────────
@pytest.mark.asyncio
async def test_scheduler_processes_programada_campaign():
    """Insert a campaign with status='programada' and past scheduled_at,
    then invoke _process_scheduled_campaigns and verify it transitions."""
    import sys
    sys.path.insert(0, "/app/backend")
    from database import connect_to_mongo, get_database
    from scheduler import _process_scheduled_campaigns

    await connect_to_mongo()
    db = await get_database()
    # Find any company_id from existing tenants
    comp = await db.companies.find_one({}, {"_id": 0, "id": 1})
    assert comp, "no company found in DB to attach test campaign"
    company_id = comp["id"]

    past = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
    camp_id = f"TEST_iter60_sched_{uuid.uuid4().hex[:8]}"
    camp = {
        "id": camp_id,
        "company_id": company_id,
        "name": "TEST_iter60_scheduled",
        "status": "programada",
        "scheduled_at": past,
        "messages": ["Oi"],
        "audiencia": [{"name": "Tester", "phone": "5511900000077"}],
        "anti_block": {"enabled": False, "daily_limit": 5},
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.campaigns.insert_one(camp)
    try:
        await _process_scheduled_campaigns(db)
        # Give runner a moment
        await asyncio.sleep(0.5)
        doc = await db.campaigns.find_one({"id": camp_id}, {"_id": 0})
        assert doc is not None
        # Should NOT be 'programada' anymore — either em_execucao, concluida, or cancelada
        assert doc["status"] in ("em_execucao", "concluida", "cancelada"), (
            f"campaign still in {doc['status']} after scheduler tick"
        )
        # If em_execucao or concluida, deliveries should be seeded
        if doc["status"] in ("em_execucao", "concluida"):
            count = await db.campaign_deliveries.count_documents({"campaign_id": camp_id})
            assert count >= 1, "campaign_deliveries not seeded by scheduler"
    finally:
        await db.campaigns.delete_one({"id": camp_id})
        await db.campaign_deliveries.delete_many({"campaign_id": camp_id})
